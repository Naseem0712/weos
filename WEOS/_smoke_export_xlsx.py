"""Customer Excel must mirror quote PDF: A4, formulas, no factory BOM."""
from __future__ import annotations

import io

from openpyxl import load_workbook

from WEOS.factory.export_xlsx import export_advance_xlsx, export_ledger_xlsx, export_quote_xlsx

A4 = 9


def _ok(cond: bool, msg: str) -> None:
    if not cond:
        raise SystemExit(f"FAIL: {msg}")
    print("OK:", msg)


def main() -> None:
    company = {
        "companyName": "ALLUKRAFT WINDOWS",
        "address": "Factory Road, City",
        "gstNo": "29ABCDE1234F1Z5",
        "phone": "9999999999",
        "email": "hi@example.com",
        "terms": "1. Valid 15 days.\n2. GST extra.",
    }
    lines = [
        {
            "displayName": "Sliding window",
            "productType": "sliding",
            "product": "29mm_sliding",
            "width": 1200,
            "height": 1500,
            "qty": 2,
            "sellingRate": 800,
            "saleUnit": "sqft",
            "glass": "5mm clear toughened",
            "colour": "white",
            "locationName": "Master Bedroom",
            "selling": {"sellingRate": 800, "saleUnit": "sqft", "sellingAmount": 3100.0},
            "commercialTotal": 3100.0,
        },
        {
            "displayName": "Door",
            "productType": "door",
            "product": "style_door",
            "width": 900,
            "height": 2100,
            "qty": 1,
            "sellingRate": 12000,
            "saleUnit": "opening",
            "locationName": "Main",
            "selling": {"sellingRate": 12000, "saleUnit": "opening", "sellingAmount": 12000.0},
            "commercialTotal": 12000.0,
        },
    ]
    payload = {
        "quotationId": "AK-1001",
        "customer": "Test Customer",
        "name": "Villa",
        "quoteDate": "2026-08-13",
        "description": "Front elevation",
        "terms": company["terms"],
        "lines": lines,
        "customerProfile": {"phone": "8888888888", "address": "Site 1", "gstNo": "29CUST1234A1Z1"},
        "company": company,
        "price": {"total": 15100},
    }
    ledger = {
        "customer": "Test Customer",
        "profile": {"name": "Test Customer", "phone": "8888888888", "address": "Site 1"},
        "asOf": "2026-08-13T10:00:00+00:00",
        "projects": [
            {
                "quotationId": "AK-1001",
                "name": "Villa",
                "version": 1,
                "status": "draft",
                "totalTaxable": 15100,
                "totalGrand": 17818,
                "grandTotal": 15100,
            }
        ],
        "advances": [
            {"paidAt": "2026-08-01", "amount": 5000, "paymentMode": "upi", "quoteId": "AK-1001", "reference": "UPI-1"},
            {"paidAt": "2026-08-10", "amount": 2000, "paymentMode": "cash", "quoteId": "AK-1001", "note": "cash"},
        ],
        "totals": {"totalTaxable": 15100, "totalGrand": 17818, "totalAdvances": 7000, "balance": 8100},
    }

    raw = export_quote_xlsx(payload, company, ledger=ledger)
    _ok(len(raw) > 2000, f"quote xlsx bytes {len(raw)}")
    wb = load_workbook(io.BytesIO(raw))
    _ok("Quote" in wb.sheetnames, f"quote sheet {wb.sheetnames}")
    _ok("Ledger" in wb.sheetnames, "ledger sheet present")
    ws = wb["Quote"]
    _ok(ws.page_setup.paperSize == A4, f"quote A4 paperSize={ws.page_setup.paperSize}")
    _ok((ws.page_setup.orientation or "portrait").lower() == "portrait", f"quote orientation {ws.page_setup.orientation}")
    _ok(int(ws.page_setup.fitToWidth or 0) == 1, f"fitToWidth={ws.page_setup.fitToWidth}")
    blob = " ".join(str(c.value or "") for row in ws.iter_rows(max_row=80, max_col=8) for c in row).lower()
    _ok("allukraft" in blob, "company letterhead")
    _ok("ak-1001" in blob, "quote number")
    _ok("test customer" in blob, "customer")
    _ok("master bedroom" in blob, "location")
    _ok("gst" in blob, "GST row")
    _ok("grand total" in blob, "grand total")
    _ok("terms" in blob, "terms section")
    _ok("bom" not in blob and "purchase rate" not in blob and "factory cost" not in blob, "no factory BOM")

    formulas = [str(c.value) for row in ws.iter_rows(max_row=120, max_col=13) for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    _ok(any("ROUND(E" in f and "*K" in f and "*F" in f for f in formulas), f"line amount qty*unit*rate {formulas[:8]}")
    _ok(any("SUMIF" in f for f in formulas), "type subtotal SUMIF")
    _ok(any("/100" in f or "Gst" in f or "G" in f and "*G" in f for f in formulas), "GST formula")
    _ok(any("SUM(G" in f for f in formulas), "taxable SUM of amounts")

    ls = wb["Ledger"]
    _ok(ls.page_setup.paperSize == A4, f"ledger A4 paperSize={ls.page_setup.paperSize}")
    led_f = [str(c.value) for row in ls.iter_rows(max_row=80, max_col=7) for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    _ok(any("SUM(B" in f for f in led_f), f"advance SUM {led_f}")
    _ok(any("-B" in f or "- B" in f for f in led_f), f"balance grand-advances {led_f}")

    lraw = export_ledger_xlsx(ledger, company)
    lwb = load_workbook(io.BytesIO(lraw))
    _ok(lwb.active.page_setup.paperSize == A4, "standalone ledger A4")
    _ok(any(isinstance(c.value, str) and c.value.startswith("=SUM") for row in lwb.active.iter_rows(max_row=80, max_col=6) for c in row), "standalone ledger formulas")

    sraw = export_advance_xlsx(
        {"id": 12, "amount": 5000, "paymentMode": "upi", "paidAt": "2026-08-01", "quoteId": "AK-1001", "reference": "UPI-1"},
        company=company,
        ledger=ledger,
        customer="Test Customer",
    )
    swb = load_workbook(io.BytesIO(sraw))
    ss = swb.active
    _ok(ss.page_setup.paperSize == A4, "slip A4")
    _ok((ss.page_setup.orientation or "portrait").lower() == "portrait", "slip portrait")
    slip_f = [str(c.value) for row in ss.iter_rows(max_row=80, max_col=5) for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    _ok(any("SUM(B" in f for f in slip_f), f"slip advance sum {slip_f}")
    _ok(any("-B" in f for f in slip_f), "slip balance formula")
    print("OK export xlsx A4 + formulas smoke passed")


if __name__ == "__main__":
    main()
