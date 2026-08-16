"""Import a running project from Excel / PDF (multi-stage quotes + advances).

WoodenMax-style workbooks (Numbers export) have one sheet per stage plus an
accounts sheet. PDF is a fallback when tables can be read. The operator fills
customer name / mobile / address; extracted lines stay editable after save.
"""

from __future__ import annotations

import hashlib
import logging
import re
from datetime import datetime, timezone
from typing import Any, Mapping

from WEOS.factory.package_quote import merge_package_quotes, normalize_package_quotes

_log = logging.getLogger("weos.project_import")

_SKIP_SHEET = re.compile(r"(?i)export\s*summary|glass\s*size|^sheet\s*1")
_ACCOUNT_SHEET = re.compile(r"(?i)account|payment")
_TOTAL_ROW = re.compile(r"(?i)^(total|g\.?\s*total|grand\s*total|value|balance|due)\b")
_GST_ROW = re.compile(r"(?i)\bgst\b")
_MONEY = re.compile(r"(?:₹|rs\.?)\s*([\d,]+(?:\.\d+)?)", re.I)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _txt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).replace("\xa0", " ").strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = _txt(v).replace(",", "").replace("₹", "").replace("Rs.", "").replace("rs", "").strip()
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", s):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _money(v: Any) -> float:
    n = _num(v)
    return round(n, 2) if n is not None else 0.0


def _iso_date(v: Any) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = _txt(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y", "%b %d %Y"):
        try:
            return datetime.strptime(s[:20].strip(), fmt).date().isoformat()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d).date().isoformat()
        except ValueError:
            return None
    return None


def _guess_category(text: str) -> str:
    t = (text or "").lower()
    if re.search(r"vent|ventilator", t):
        return "ventilator"
    if re.search(r"rail", t):
        return "railing"
    if re.search(r"grill", t):
        return "grill"
    if re.search(r"\bgate\b|iron gate", t):
        return "gate"
    if re.search(r"kg\b|iron leader|fabrication", t):
        return "iron_fabrication"
    if re.search(r"casement|top hunk", t):
        return "casement"
    if re.search(r"sliding|window|fold", t):
        return "window"
    if re.search(r"louver", t):
        return "louver"
    return "other"


def _guess_unit(headers: str, text: str, extra: str = "") -> str:
    blob = f"{headers} {text} {extra}".lower()
    if "rft" in blob or re.search(r"\brft\b", blob):
        return "rft"
    if re.search(r"\bkg\b", blob):
        return "kg"
    if "sft" in blob or "area" in blob:
        return "sft"
    return "pcs"


def _pay_mode(raw: Any) -> str:
    t = _txt(raw).lower()
    if "gpay" in t or "g.pay" in t or "upi" in t:
        return "upi"
    if "neft" in t:
        return "neft"
    if "rtgs" in t:
        return "rtgs"
    if "cheque" in t or "check" in t:
        return "cheque"
    if "card" in t:
        return "card"
    if "transfer" in t or "tranfer" in t or "neft" in t:
        return "neft"
    if "cash" in t:
        return "cash"
    return "other" if t else "cash"


def _header_map(row: list[Any]) -> dict[str, int] | None:
    labels = [_txt(c).lower() for c in row]
    joined = " ".join(labels)
    if "amount" not in joined:
        return None
    if not any(k in joined for k in ("rate", "width", "height", "length", "count", "qty", "rft", "sft")):
        return None
    if "firmat" in joined or "format" in joined:
        return None
    idx: dict[str, int] = {}
    for i, lab in enumerate(labels):
        if lab in {"amount", "amt", "value"} and "amount" not in idx:
            idx["amount"] = i
        elif lab in {"rate", "rs", "₹"}:
            idx["rate"] = i
        elif lab in {"width", "w", "length", "depth"} or lab.startswith("length"):
            idx.setdefault("width", i)
        elif lab in {"height", "h"} or lab.startswith("height"):
            idx["height"] = i
        elif lab in {"count", "qty", "nos", "no"}:
            idx["qty"] = i
        elif "rft" in lab:
            idx["area"] = i
            idx["unit_rft"] = 1
        elif lab in {"area", "sft", "sqft"} or "sft" in lab or "area" in lab:
            idx.setdefault("area", i)
        elif any(k in lab for k in ("desc", "product", "item")):
            idx["desc"] = i
        elif lab in {"name"}:
            idx.setdefault("loc", i)
    if "amount" not in idx:
        return None
    return idx


def _row_cells(row: Any) -> list[Any]:
    if isinstance(row, (list, tuple)):
        return list(row)
    return []


_SKIP_ITEM = re.compile(
    r"(?i)payment\s*term|client\s*/\s*company|for woodenmax|"
    r"^(page\s+\d+|advance|due|name\s*:|a/c\s*:|gst\s*bill)\b|"
    r"ifsc|hdfc|branch\s*:"
)


def _item_from_row(cells: list[Any], headers: str = "", last_note: str = "") -> dict[str, Any] | None:
    texts = [_txt(c) for c in cells]
    blob = " ".join(t for t in texts if t).strip()
    if not blob or _TOTAL_ROW.match(blob) or _GST_ROW.search(blob[:40]):
        return None
    if _SKIP_ITEM.search(blob):
        return None
    unit_hint = ""
    for t in texts:
        if t.lower() in {"kg", "no", "nos", "pc", "pcs", "rft", "sft", "mtr"}:
            unit_hint = t.lower()
    loc = ""
    desc = ""
    rest = list(cells)
    first_n = _num(cells[0]) if cells else None
    second = _txt(cells[1]) if len(cells) > 1 else ""
    if (
        first_n is not None
        and abs(first_n - round(first_n)) < 1e-6
        and 1 <= first_n <= 80
        and second
        and _num(cells[1]) is None
    ):
        desc = second
        rest = list(cells[2:])
    elif cells and _num(cells[0]) is None:
        loc = texts[0] if texts[0] and len(texts[0]) < 28 else ""
        if len(cells) > 1 and _num(cells[1]) is None and texts[1]:
            desc = texts[1]
            rest = list(cells[2:])
        else:
            desc = loc
            loc = ""
            rest = list(cells[1:])
    nums = [n for n in (_num(c) for c in rest) if n is not None]
    if not nums:
        return None
    amt = nums[-1]
    if amt is None or amt <= 0:
        return None
    rate = nums[-2] if len(nums) >= 2 else None
    qty = None
    w = h = None
    if amt < 400 and rate is not None and rate < 10:
        return None
    if len(nums) >= 6:
        w, h, qty, _area, rate, amt = nums[-6:]
    elif len(nums) == 5:
        a, b, c, d, e = nums
        if d and abs(c * d - e) <= max(1.0, 0.03 * e):
            w, h, qty, rate, amt = a, b, 1.0, d, e
        elif d and abs(a * d - e) <= max(1.0, 0.03 * e):
            qty, rate, amt = a, d, e
        else:
            w, h, qty, rate, amt = a, b, c, d, e
    elif len(nums) == 4:
        if unit_hint:
            qty, rate, amt = nums[-3], nums[-2], nums[-1]
        else:
            a, b, c, d = nums
            if c and abs(a * c - d) <= max(1.0, 0.03 * d):
                qty, rate, amt = a, c, d
            else:
                qty, rate, amt = b, c, d
    elif len(nums) == 3:
        qty, rate, amt = nums
    elif len(nums) == 2:
        rate, amt = nums
    if not desc:
        desc = last_note
    if re.match(r"(?i)^(total|g\.?\s*total|grand\s*total)$", desc.strip()):
        return None
    if loc and desc and loc.lower() not in desc.lower():
        desc = f"{loc} — {desc}"
    elif loc and not desc:
        desc = loc
    if not desc:
        return None
    size = None
    if w and h:
        size = f"{w:g}x{h:g}"
    elif w:
        size = f"{w:g}"
    unit = _guess_unit(headers, desc, unit_hint)
    return {
        "category": _guess_category(desc),
        "qty": qty if qty and qty > 0 else None,
        "size": size,
        "unit": unit,
        "amount": round(float(amt), 2),
        "note": desc[:240],
        "rate": round(float(rate), 2) if rate else None,
    }


def _sheet_title(rows: list[list[Any]]) -> tuple[str, str | None, str | None]:
    title = ""
    quote_no = None
    dated = None
    for cells in rows[:16]:
        if _header_map(cells):
            break
        blob = " ".join(_txt(c) for c in cells if _txt(c))
        if not blob:
            continue
        nums = [n for n in (_num(c) for c in cells) if n is not None]
        m = re.search(r"(?i)P\.?\s*I\.?\s*Number\s*[:\-]*\s*([A-Za-z0-9][\w./-]*)", blob)
        if m:
            quote_no = m.group(1).strip()
        m = re.search(r"(?i)TITEL\s*[:\-]*\s*(.+?)(?:P\.?\s*I\.?|$)", blob)
        if m and not title:
            title = m.group(1).strip(" .")
        for c in cells:
            if isinstance(c, datetime) and not dated:
                dated = c.date().isoformat()
        if title or len(nums) >= 3:
            continue
        if re.search(r"(?i)^(railings?|cladding|bathrooms?|entrance|grills?|acp|hpl|office|luxury windows?)\b", blob.strip()) and len(blob) < 90:
            title = blob.split("\n")[0].strip()
        elif re.search(r"(?i)(safety grills|cladding|showers|entrance doors|railings)", blob) and 8 < len(blob) < 120:
            if "woodenmax" not in blob.lower() and "gstin" not in blob.lower():
                title = blob.split("\n")[0].strip()
    return title, quote_no, dated


def _customer_from_rows(rows: list[list[Any]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for cells in rows[:20]:
        blob = " ".join(_txt(c) for c in cells if _txt(c))
        m = re.search(r"(?i)CLINT\s*/\s*COMPANY\s*[:-]\s*(.+?)(?:GST|ADDRESS|$)", blob)
        if m:
            out.setdefault("name", m.group(1).strip(" ."))
        m = re.search(r"(?i)ADDRESS\s*[:-]\s*(.+?)(?:TITEL|$)", blob)
        if m:
            out.setdefault("address", m.group(1).strip(" ."))
        if re.match(r"(?i)to$", _txt(cells[0]) if cells else ""):
            continue
        if re.search(r"(?i)rajkumar|west margpally", blob) and "woodenmax" not in blob.lower():
            out.setdefault("name", re.sub(r"(?i)^to\s*", "", blob.split("\n")[0]).strip(" ,"))
            if "hyderabad" in blob.lower() or "margpally" in blob.lower():
                out.setdefault("address", blob.split("\n")[0].strip())
    return out


def parse_account_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    advances: list[dict[str, Any]] = []
    header_i = None
    for i, cells in enumerate(rows):
        labs = " ".join(_txt(c).lower() for c in cells)
        if "amount" in labs and ("name" in labs or "firmat" in labs or "format" in labs or "date" in labs):
            header_i = i
            break
    start = (header_i + 1) if header_i is not None else 0
    for cells in rows[start:]:
        blob = " ".join(t for t in (_txt(c) for c in cells) if t).strip()
        if _TOTAL_ROW.match(blob) or re.search(r"(?i)^(total|value|balance|due)\b", blob):
            continue
        texts = [_txt(c) for c in cells]
        nums = [_num(c) for c in cells]
        amt = None
        for n in nums:
            if n is not None and n >= 100:
                amt = n
                break
        if not amt:
            continue
        purpose = ""
        mode_raw = ""
        dated = None
        for i, c in enumerate(cells):
            t = _txt(c)
            if not t:
                continue
            if _num(c) == amt:
                continue
            if isinstance(c, datetime) or _iso_date(t):
                dated = dated or (c.date().isoformat() if isinstance(c, datetime) else _iso_date(t))
                continue
            if re.search(r"(?i)cash|transfer|tranfer|gpay|g\.pay|upi|neft|cheque", t) and len(t) < 24:
                mode_raw = t
                continue
            if re.fullmatch(r"\d{1,3}", t):
                continue
            if not purpose:
                purpose = t
        if re.match(r"(?i)^(total|value|balance|due)$", purpose.strip()):
            continue
        if amt >= 100:
            advances.append(
                {
                    "amount": round(float(amt), 2),
                    "paymentMode": _pay_mode(mode_raw or purpose),
                    "paidAt": dated,
                    "purpose": purpose.strip(),
                    "note": purpose.strip() or None,
                    "reference": purpose.strip() or None,
                }
            )
    return advances


def parse_quote_sheet(rows: list[list[Any]], *, sheet_name: str = "") -> dict[str, Any] | None:
    header_i = None
    header_join = sheet_name
    for i, cells in enumerate(rows):
        hm = _header_map(cells)
        if hm:
            header_i = i
            header_join = " ".join(_txt(c) for c in cells)
            break
    if header_i is None:
        return None
    items: list[dict[str, Any]] = []
    gst_mode = "off"
    gst_percent = 0.0
    gst_amount = None
    grand = None
    taxable = None
    saw_total = False
    last_note = ""
    for cells in rows[header_i + 1 :]:
        blob = " ".join(_txt(c) for c in cells)
        if not blob:
            continue
        if _GST_ROW.search(blob) and any(_num(c) is not None for c in cells):
            nums = [n for n in (_num(c) for c in cells) if n is not None]
            if nums:
                gst_amount = round(nums[-1], 2)
            pct = None
            for n in nums[:-1] if len(nums) > 1 else []:
                if 0 < n <= 1.5:
                    pct = n * 100.0 if n <= 1 else n
                elif 1 < n <= 28:
                    pct = n
            gst_mode = "exclude"
            if pct:
                gst_percent = pct
            continue
        if re.search(r"(?i)g\.?\s*total|grand\s*total", blob):
            nums = [n for n in (_num(c) for c in cells) if n and n > 10]
            if nums:
                grand = round(nums[-1], 2)
            continue
        if re.search(r"(?i)\btotal\b", blob):
            nums = [n for n in (_num(c) for c in cells) if n and n > 100]
            if nums:
                taxable = round(nums[-1], 2)
                saw_total = True
            continue
        it = _item_from_row(cells, header_join, last_note=last_note)
        if it:
            items.append(it)
            last_note = it.get("note") or last_note
    if not items:
        return None
    title, qno, dated = _sheet_title(rows)
    sheet_label = re.sub(r"\s*-\s*35 mm.*", "", sheet_name, flags=re.I).strip(" -") or sheet_name
    low = (title or "").lower()
    if sheet_label and (
        not title
        or low.startswith("entrance doors")
        or low.startswith("railings")
        or low in {"to", "glass railing"}
    ):
        title = sheet_label
    cust = _customer_from_rows(rows)
    if taxable is None:
        taxable = round(sum(it["amount"] for it in items), 2)
    if gst_mode == "exclude" and gst_amount is None and gst_percent:
        gst_amount = round(taxable * gst_percent / 100.0, 2)
    if grand is None:
        grand = round(taxable + (gst_amount or 0), 2)
    complete = bool(saw_total or (gst_amount is not None))
    return {
        "title": title,
        "quotationId": qno,
        "quoteDate": dated,
        "customerHint": cust,
        "gstMode": gst_mode if (gst_amount or gst_percent) else "exclude",
        "gstPercent": gst_percent,
        "gstAmount": gst_amount,
        "totalTaxable": taxable,
        "projectValue": grand,
        "totalGrand": grand,
        "note": title,
        "items": items,
        "sheetName": sheet_name,
        "complete": complete,
    }


def parse_excel_bytes(data: bytes, filename: str = "project.xlsx") -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl required for Excel import") from exc
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    quotes: list[dict[str, Any]] = []
    advances: list[dict[str, Any]] = []
    customer: dict[str, str] = {}
    pending: dict[str, Any] | None = None
    for name in wb.sheetnames:
        if _SKIP_SHEET.search(name or ""):
            continue
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if _ACCOUNT_SHEET.search(name or "") or any(
            "firmat" in " ".join(_txt(c).lower() for c in row) or "format" in " ".join(_txt(c).lower() for c in row)
            for row in rows[:6]
        ):
            if any("amount" in " ".join(_txt(c).lower() for c in row) and "name" in " ".join(_txt(c).lower() for c in row) for row in rows[:8]):
                advances.extend(parse_account_rows(rows))
                customer.update(_customer_from_rows(rows))
                continue
        q = parse_quote_sheet(rows, sheet_name=name)
        if not q:
            continue
        customer.update(q.get("customerHint") or {})
        page_text = " ".join(_txt(c) for row in rows[:12] for c in row)
        is_cont = bool(re.search(r"(?i)page\s*2", page_text))
        if pending and is_cont:
            _absorb_quote(pending, q)
            if pending.get("complete"):
                quotes.append(pending)
                pending = None
            continue
        if pending:
            quotes.append(pending)
        pending = q
    if pending:
        quotes.append(pending)
    return _pack(quotes, advances, customer, sources=[filename], file_sha256=_file_sha256(data))


def parse_pdf_bytes(data: bytes, filename: str = "project.pdf") -> dict[str, Any]:
    from WEOS.paths import output_dir

    tmp = output_dir() / "_import_tmp.pdf"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(data)
    try:
        from WEOS.learning.pdf_catalogue import _read_pdf_pages

        pages = _read_pdf_pages(tmp)
    except Exception:
        pages = []
    finally:
        tmp.unlink(missing_ok=True)
    quotes: list[dict[str, Any]] = []
    advances: list[dict[str, Any]] = []
    customer: dict[str, str] = {}
    pending: dict[str, Any] | None = None
    for page in pages:
        tables = page.get("tables") or []
        text = page.get("text") or ""
        rows: list[list[Any]] = []
        for t in tables:
            for r in t or []:
                rows.append(list(r or []))
        if not rows and text:
            rows = [[ln] for ln in text.splitlines() if ln.strip()]
        low = text.lower()
        if "firmat" in low or (re.search(r"(?i)rajkumar ji accounts", text) and "amount" in low):
            advances.extend(parse_account_rows(rows if rows else [[ln] for ln in text.splitlines()]))
            # Fallback: regex ledger lines from raw text.
            if not advances:
                for ln in text.splitlines():
                    m = re.search(r"(?i)(\d+)\s+(.+?)\s+([\d,]+)\s+(cash|transfer|tranfer|gpay|g\.pay|upi)", ln)
                    if m:
                        advances.append(
                            {
                                "amount": _money(m.group(3)),
                                "paymentMode": _pay_mode(m.group(4)),
                                "purpose": m.group(2).strip(),
                                "note": m.group(2).strip(),
                                "reference": m.group(2).strip(),
                                "paidAt": _iso_date(ln),
                            }
                        )
            customer.update(_customer_from_rows(rows))
            continue
        q = parse_quote_sheet(rows, sheet_name=f"page {page.get('page')}")
        if not q:
            continue
        customer.update(q.get("customerHint") or {})
        if pending and not pending.get("complete") and not q.get("complete"):
            _absorb_quote(pending, q)
            if pending.get("complete"):
                quotes.append(pending)
                pending = None
            continue
        if pending:
            quotes.append(pending)
        pending = q
    if pending:
        quotes.append(pending)
    return _pack(quotes, advances, customer, sources=[filename], file_sha256=_file_sha256(data))


def _absorb_quote(pending: dict[str, Any], nxt: dict[str, Any]) -> None:
    pending["items"].extend(nxt.get("items") or [])
    for k in ("totalTaxable", "gstAmount", "projectValue", "totalGrand", "gstMode", "gstPercent"):
        if nxt.get(k) not in (None, "", 0, 0.0):
            pending[k] = nxt[k]
    if nxt.get("complete"):
        pending["complete"] = True
    if nxt.get("quotationId") and not pending.get("quotationId"):
        pending["quotationId"] = nxt["quotationId"]


def _file_sha256(data: bytes) -> str:
    return hashlib.sha256(data or b"").hexdigest()


def _pack(
    quotes: list[dict[str, Any]],
    advances: list[dict[str, Any]],
    customer: dict[str, str],
    *,
    sources: list[str],
    file_sha256: str | None = None,
) -> dict[str, Any]:
    src_name = (sources or [None])[0]
    payload = []
    for i, q in enumerate(quotes):
        payload.append(
            {
                "id": f"pq_imp_{i+1}",
                "quotationId": q.get("quotationId") if str(q.get("quotationId") or "").strip() not in {"", "-", "—"} else None,
                "note": q.get("note") or q.get("title"),
                "gstMode": q.get("gstMode") or "exclude",
                "gstPercent": q.get("gstPercent") or 0,
                "gstAmount": q.get("gstAmount"),
                "projectValue": q.get("projectValue"),
                "totalGrand": q.get("totalGrand"),
                "items": q.get("items") or [],
                "sheetName": q.get("sheetName"),
                "sourceFile": q.get("sourceFile") or src_name,
                "sourceFileSha256": q.get("sourceFileSha256") or file_sha256,
            }
        )
    norm = normalize_package_quotes(payload)
    adv_total = round(sum(_money(a.get("amount")) for a in advances), 2)
    value = round(sum(_money(q.get("projectValue") or q.get("totalGrand")) for q in norm), 2)
    return {
        "ok": True,
        "customerHint": {
            "name": customer.get("name") or "",
            "address": customer.get("address") or "",
        },
        "quotes": norm,
        "quoteCount": len(norm),
        "advances": advances,
        "advanceCount": len(advances),
        "advanceTotal": adv_total,
        "projectValue": value,
        "balance": round(value - adv_total, 2),
        "sources": sources,
        "fileSha256": file_sha256,
        "stages": [
            {
                "title": q.get("note") or q.get("sheetName") or q.get("quotationId"),
                "sheetName": q.get("sheetName"),
                "items": len(q.get("items") or []),
                "value": q.get("projectValue"),
                "fingerprint": q.get("importFingerprint"),
            }
            for q in norm
        ],
    }


def merge_previews(*packs: Mapping[str, Any]) -> dict[str, Any]:
    excel_packs = []
    other_packs = []
    for p in packs:
        if not p:
            continue
        srcs = " ".join(str(s).lower() for s in (p.get("sources") or []))
        if any(srcs.endswith(ext) or ext in srcs for ext in (".xlsx", ".xlsm", ".xls")):
            excel_packs.append(p)
        else:
            other_packs.append(p)
    preferred = excel_packs or other_packs
    fallback = other_packs if excel_packs else []
    quotes: list[Any] = []
    advances: list[Any] = []
    customer: dict[str, str] = {}
    sources: list[str] = []
    for p in preferred:
        quotes.extend(list(p.get("quotes") or []))
        advances.extend(list(p.get("advances") or []))
        customer.update(p.get("customerHint") or {})
        sources.extend(list(p.get("sources") or []))
    for p in fallback:
        if not advances:
            advances.extend(list(p.get("advances") or []))
        if not quotes:
            quotes.extend(list(p.get("quotes") or []))
        customer.update(p.get("customerHint") or {})
        sources.extend(list(p.get("sources") or []))
    return _pack(
        [
            {
                "quotationId": q.get("quotationId"),
                "title": q.get("note"),
                "note": q.get("note"),
                "gstMode": q.get("gstMode"),
                "gstPercent": q.get("gstPercent"),
                "gstAmount": q.get("gstAmount"),
                "projectValue": q.get("projectValue"),
                "totalGrand": q.get("totalGrand"),
                "items": q.get("items") or [],
                "sheetName": q.get("sheetName"),
                "sourceFile": q.get("sourceFile"),
                "sourceFileSha256": q.get("sourceFileSha256"),
            }
            for q in quotes
        ],
        advances,
        customer,
        sources=sources,
        file_sha256=next((str(p.get("fileSha256") or "") for p in preferred if p.get("fileSha256")), None) or None,
    )


def parse_upload(filename: str, data: bytes) -> dict[str, Any]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return parse_excel_bytes(data, filename)
    if name.endswith(".pdf"):
        return parse_pdf_bytes(data, filename)
    raise ValueError("Upload an Excel (.xlsx) or PDF of the project")


def _advance_is_duplicate(existing: list[Mapping[str, Any]], incoming: Mapping[str, Any]) -> bool:
    try:
        amt = round(float(incoming.get("amount") or 0), 2)
    except (TypeError, ValueError):
        return False
    paid = str(incoming.get("paidAt") or "")[:10]
    purpose = re.sub(
        r"\s+",
        " ",
        str(incoming.get("purpose") or incoming.get("reference") or incoming.get("note") or "").strip().lower(),
    )
    mode = str(incoming.get("paymentMode") or "").strip().lower()
    for a in existing:
        try:
            if abs(float(a.get("amount") or 0) - amt) > 0.51:
                continue
        except (TypeError, ValueError):
            continue
        a_paid = str(a.get("paidAt") or a.get("paid_at") or "")[:10]
        a_purp = re.sub(r"\s+", " ", str(a.get("reference") or a.get("note") or "").strip().lower())
        a_mode = str(a.get("paymentMode") or a.get("payment_mode") or "").strip().lower()
        if paid and a_paid and paid != a_paid:
            continue
        if purpose and a_purp and purpose != a_purp:
            continue
        if mode and a_mode and mode not in {a_mode, "other"} and a_mode not in {mode, "other"} and paid:
            continue
        return True
    return False


def plan_import_merge(existing_quotes: list[Any] | None, incoming_quotes: list[Any] | None, *, project_id: str | None = None) -> dict[str, Any]:
    merged = merge_package_quotes(existing_quotes or [], incoming_quotes or [], project_id=project_id)
    return {
        "added": merged.get("added") or [],
        "skipped": merged.get("skipped") or [],
        "updated": merged.get("updated") or [],
        "addedCount": merged.get("addedCount") or 0,
        "skippedCount": merged.get("skippedCount") or 0,
        "updatedCount": merged.get("updatedCount") or 0,
        "quoteCountAfter": merged.get("quoteCount") or 0,
        "quotes": merged.get("quotes") or [],
    }


def commit_imported_project(
    preview: Mapping[str, Any],
    *,
    customer: str,
    customer_mobile: str | None = None,
    customer_address: str | None = None,
    customer_gst: str | None = None,
    project_name: str | None = None,
    quotation_id: str | None = None,
    project_id: str | None = None,
    company_gst: str | None = None,
    import_advances: bool = True,
) -> dict[str, Any]:
    """Save extracted stages onto a WEOS project and optional customer ledger."""
    from WEOS.factory.customer_store import save_customer_profile
    from WEOS.factory.ledger_store import add_advance, list_advances_for_projects
    from WEOS.factory.project_store import empty_project, load_project, save_project

    name = (customer or (preview.get("customerHint") or {}).get("name") or "").strip()
    if not name:
        raise ValueError("Enter the customer name so this project and account can be saved")
    quotes = list(preview.get("quotes") or [])
    if not quotes:
        raise ValueError("No quote stages found in the file — check Excel/PDF and try again")

    profile = {
        "name": name,
        "phone": (customer_mobile or "").strip() or None,
        "address": (customer_address or (preview.get("customerHint") or {}).get("address") or "").strip() or None,
        "gstNo": (customer_gst or "").strip() or None,
        "companyGst": company_gst,
    }
    save_customer_profile(name, {k: v for k, v in profile.items() if v})

    existing: list[Any] = []
    if project_id:
        doc = load_project(str(project_id).strip())
        existing = list(doc.get("packageQuotes") or [])
    else:
        doc = empty_project(name=project_name or f"{name} project", customer=name)
    merged = merge_package_quotes(existing, quotes, project_id=str(doc.get("projectId") or project_id or "") or None)
    if not merged.get("quotes") and not existing:
        raise ValueError("No quote stages found in the file — check Excel/PDF and try again")
    doc["customer"] = name
    doc["customerMobile"] = (customer_mobile or "").strip() or doc.get("customerMobile")
    doc["customerAddress"] = (customer_address or profile.get("address") or "") or doc.get("customerAddress")
    doc["customerGst"] = (customer_gst or "").strip() or doc.get("customerGst")
    if company_gst:
        doc["companyGst"] = company_gst
    first_qid = quotation_id or next((q.get("quotationId") for q in (merged.get("quotes") or quotes) if q.get("quotationId")), None)
    if first_qid and not doc.get("quotationId"):
        doc["quotationId"] = str(first_qid)
    doc["packageQuotes"] = merged.get("quotes") or existing
    doc["quoteKind"] = "package" if not (doc.get("lines") or []) else "mixed"
    log = list(doc.get("revisionLog") or [])
    log.append(
        {
            "at": _now(),
            "action": "import_project_files",
            "sources": list(preview.get("sources") or []),
            "quoteCount": len(doc.get("packageQuotes") or []),
            "addedCount": merged.get("addedCount") or 0,
            "skippedCount": merged.get("skippedCount") or 0,
            "updatedCount": merged.get("updatedCount") or 0,
            "advanceCount": len(preview.get("advances") or []),
        }
    )
    doc["revisionLog"] = log[-80:]
    prev_meta = doc.get("importMeta") if isinstance(doc.get("importMeta"), dict) else {}
    hashes = list(prev_meta.get("fileHashes") or [])
    sha = str(preview.get("fileSha256") or "").strip()
    if sha and sha not in hashes:
        hashes.append(sha)
    doc["importMeta"] = {
        "at": _now(),
        "sources": list(dict.fromkeys(list(prev_meta.get("sources") or []) + list(preview.get("sources") or []))),
        "stages": preview.get("stages") or [],
        "fileHashes": hashes[-12:],
        "lastMerge": {
            "added": merged.get("added") or [],
            "skipped": merged.get("skipped") or [],
            "updated": merged.get("updated") or [],
        },
    }
    saved = save_project(doc, bump_version=True, action="import_project_files")
    imported_adv = []
    skipped_adv = 0
    if import_advances:
        existing_adv = list_advances_for_projects([str(saved.get("projectId") or "")])
        for a in preview.get("advances") or []:
            if _advance_is_duplicate(existing_adv + imported_adv, a):
                skipped_adv += 1
                continue
            try:
                row = add_advance(
                    name,
                    {
                        "amount": a.get("amount"),
                        "paymentMode": a.get("paymentMode") or "cash",
                        "paidAt": a.get("paidAt"),
                        "note": a.get("purpose") or a.get("note"),
                        "reference": a.get("reference") or a.get("purpose"),
                        "projectId": saved.get("projectId"),
                        "quoteId": "any",
                        "customerName": name,
                        "companyGst": company_gst,
                    },
                )
                imported_adv.append(row)
            except Exception:
                _log.exception("import advance skipped")
    value = round(sum(_money(q.get("projectValue") or q.get("totalGrand")) for q in (saved.get("packageQuotes") or [])), 2)
    return {
        "ok": True,
        "project": {
            "projectId": saved.get("projectId"),
            "quotationId": saved.get("quotationId"),
            "name": saved.get("name"),
            "version": saved.get("version"),
        },
        "customer": name,
        "quoteCount": len(saved.get("packageQuotes") or []),
        "advanceCount": len(imported_adv),
        "advanceSkipped": skipped_adv,
        "addedCount": merged.get("addedCount") or 0,
        "skippedCount": merged.get("skippedCount") or 0,
        "updatedCount": merged.get("updatedCount") or 0,
        "added": merged.get("added") or [],
        "skipped": merged.get("skipped") or [],
        "updated": merged.get("updated") or [],
        "projectValue": value,
        "advanceTotal": preview.get("advanceTotal"),
        "balance": preview.get("balance"),
    }
