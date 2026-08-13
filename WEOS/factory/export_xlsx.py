"""Excel exports mirroring customer Quote / Ledger / Advance slip PDF contract.

A4 portrait, letterhead, line items + drawings, GST totals, ledger advances.
Amounts use formulas (qty × unit-qty × rate) so a user can correct cells and
re-print / PDF from Excel. Customer workbooks never include factory BOM.
"""

from __future__ import annotations

import io
import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

# openpyxl A4 paper size constant
_A4 = 9
_GST_PCT = 18.0
_SQFT_PER_SQM = 10.7639


def _txt(v: Any) -> str:
    return str(v or "").strip()


def _num(v: Any) -> float | None:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v: Any, default: int = 1) -> int:
    try:
        n = int(round(float(v if v not in (None, "") else default)))
        return n if n > 0 else default
    except (TypeError, ValueError):
        return default


def _money_fmt(cell) -> None:
    cell.number_format = "#,##0.00"


def _qty_fmt(cell) -> None:
    cell.number_format = "0"


def _unit_fmt(cell) -> None:
    cell.number_format = "0.000"


def _wb():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    styles = {
        "bold": Font(bold=True, name="Calibri", size=11),
        "title": Font(bold=True, name="Calibri", size=14, color="0A5A48"),
        "section": Font(bold=True, name="Calibri", size=12, color="0A5A48"),
        "muted": Font(name="Calibri", size=9, color="5C584F"),
        "head": Font(bold=True, name="Calibri", size=9, color="0A5A48"),
        "grand": Font(bold=True, name="Calibri", size=13, color="0A5A48"),
        "wrap": Alignment(wrap_text=True, vertical="top"),
        "right": Alignment(horizontal="right", vertical="top"),
        "fill_head": PatternFill("solid", fgColor="F7F4EE"),
        "fill_grand": PatternFill("solid", fgColor="E7F3EE"),
        "thin": Border(
            bottom=Side(style="thin", color="D1CCC4"),
        ),
    }
    return wb, styles


def _apply_a4(ws, *, last_col: str = "H", last_row: int = 40, title_rows: str | None = None) -> None:
    """Match customer quote PDF: A4 portrait, ~12 mm margins, fit-to-width 1."""
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.paperSize = _A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.55, bottom=0.55, header=0.28, footer=0.28)
    ws.print_area = f"A1:{last_col}{max(last_row, 1)}"
    if title_rows:
        ws.print_title_rows = title_rows
    ws.page_setup.paperHeight = "297mm"
    ws.page_setup.paperWidth = "210mm"


def _header_footer(ws, *, company: str, right: str, footer_left: str = "Powered by WEOS") -> None:
    ws.oddHeader.left.text = (company or "WEOS")[:60]
    ws.oddHeader.right.text = (right or "")[:60]
    ws.oddFooter.left.text = footer_left
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.center.text = ""


def _embed_png(ws, png: bytes, cell: str, *, width: int, height: int, temps: list[str]) -> None:
    from openpyxl.drawing.image import Image as XLImage

    fd, path = tempfile.mkstemp(suffix=".png")
    os.close(fd)
    Path(path).write_bytes(png)
    temps.append(path)
    img = XLImage(path)
    img.width = width
    img.height = height
    ws.add_image(img, cell)


def _logo_png_bytes(company: Mapping[str, Any] | None) -> bytes | None:
    company = company or {}
    path = _txt(company.get("logoPath"))
    if not path:
        try:
            from WEOS.factory.company_store import logo_file

            lf = logo_file()
            path = str(lf) if lf else ""
        except Exception:
            path = ""
    if not path:
        return None
    p = Path(path)
    if not p.is_file():
        return None
    suf = p.suffix.lower()
    if suf == ".svg":
        try:
            from WEOS.factory.image_engine import svg_to_png_bytes

            return svg_to_png_bytes(p.read_text(encoding="utf-8"), scale=1.2)
        except Exception:
            return None
    if suf in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
        try:
            return p.read_bytes()
        except OSError:
            return None
    return None


def _sale_unit(line: Mapping[str, Any]) -> str:
    selling = line.get("selling") if isinstance(line.get("selling"), Mapping) else {}
    opts = line.get("options") if isinstance(line.get("options"), Mapping) else {}
    rq = opts.get("railingQuote") if isinstance(opts.get("railingQuote"), Mapping) else {}
    sq = opts.get("showerQuote") if isinstance(opts.get("showerQuote"), Mapping) else {}
    vq = opts.get("ventilatorQuote") if isinstance(opts.get("ventilatorQuote"), Mapping) else {}
    raw = (
        line.get("saleUnit")
        or selling.get("saleUnit")
        or rq.get("saleUnit")
        or sq.get("saleUnit")
        or vq.get("saleUnit")
        or "sqft"
    )
    u = str(raw or "sqft").strip().lower()
    aliases = {
        "sft": "sqft",
        "sq.ft": "sqft",
        "sq.ft.": "sqft",
        "m2": "sqm",
        "m²": "sqm",
        "nos": "opening",
        "pcs": "opening",
        "each": "opening",
        "pc": "pc",
    }
    return aliases.get(u, u or "sqft")


def _unit_qty(line: Mapping[str, Any], *, qty: int, rate: float | None, amount: float | None) -> float:
    """Per-piece multiplier so Amount = Qty × UnitQty × Rate matches the quote PDF."""
    if rate and rate > 0 and amount is not None and qty > 0:
        derived = round(float(amount) / (float(qty) * float(rate)), 6)
        if derived > 0:
            return derived
    unit = _sale_unit(line)
    try:
        w = float(line.get("width") or 0)
        h = float(line.get("height") or 0)
    except (TypeError, ValueError):
        w = h = 0.0
    opts = line.get("options") if isinstance(line.get("options"), Mapping) else {}
    rq = opts.get("railingQuote") if isinstance(opts.get("railingQuote"), Mapping) else {}
    selling = line.get("selling") if isinstance(line.get("selling"), Mapping) else {}
    if unit in ("opening", "pc", "nos", "each"):
        return 1.0
    if unit == "sqm":
        return round((w * h) / 1_000_000.0, 4) or 1.0
    if unit == "rft":
        n = _num(rq.get("widthUnit") or rq.get("lengthRft") or selling.get("billableQty"))
        if n and qty:
            # billableQty may already include qty
            per = n / qty if n > qty * 2 else n
            return round(float(per), 4) or 1.0
        return round(2.0 * (w + h) / 304.8, 3) or 1.0
    if unit == "rmt":
        n = _num(rq.get("widthUnit") or rq.get("lengthRmt") or selling.get("billableQty"))
        if n and qty:
            per = n / qty if n > qty * 2 else n
            return round(float(per), 4) or 1.0
        return round(2.0 * (w + h) / 1000.0, 3) or 1.0
    area = round((w * h) / 1_000_000.0 * _SQFT_PER_SQM, 3)
    return area or 1.0


def _unit_qty_formula_or_value(line: Mapping[str, Any], row: int, unit_qty: float) -> Any:
    """Prefer live W×H→sqft formula for windows; otherwise a numeric unit qty."""
    unit = _sale_unit(line)
    try:
        w = float(line.get("width") or 0)
        h = float(line.get("height") or 0)
    except (TypeError, ValueError):
        w = h = 0.0
    if unit == "sqft" and w > 0 and h > 0:
        return f"=ROUND((I{row}/1000)*(J{row}/1000)*{_SQFT_PER_SQM},3)"
    if unit == "sqm" and w > 0 and h > 0:
        return f"=ROUND((I{row}/1000)*(J{row}/1000),4)"
    return round(float(unit_qty or 1), 6)


def _specs_text(line: Mapping[str, Any]) -> str:
    try:
        from WEOS.factory.marqt_pdf import _spec_rows

        rows = _spec_rows(line, audience="customer")
        parts: list[str] = []
        for lab, val in rows[:20]:
            lab_s, val_s = _txt(lab), _txt(val)
            if not val_s:
                continue
            parts.append(f"{lab_s}: {val_s}" if lab_s else val_s)
        if parts:
            return "\n".join(parts)
    except Exception:
        pass
    bits = [
        _txt(line.get("displayName") or line.get("productLabel")),
        f"{line.get('width')}×{line.get('height')} mm" if line.get("width") and line.get("height") else "",
        _txt(line.get("glass")),
        _txt(line.get("colour") or line.get("color")),
    ]
    return " · ".join(b for b in bits if b) or "—"


def _line_rate(line: Mapping[str, Any]) -> float | None:
    selling = line.get("selling") if isinstance(line.get("selling"), Mapping) else {}
    opts = line.get("options") if isinstance(line.get("options"), Mapping) else {}
    rq = opts.get("railingQuote") if isinstance(opts.get("railingQuote"), Mapping) else {}
    for src in (selling, line, rq, line.get("price") if isinstance(line.get("price"), Mapping) else {}):
        if not isinstance(src, Mapping):
            continue
        for key in ("sellingRate", "sellingPerUnit", "customerRate", "unitRate"):
            n = _num(src.get(key))
            if n is not None:
                return n
    amt = None
    try:
        from WEOS.factory.customer_line_view import customer_line_amount

        amt = customer_line_amount(line)
    except Exception:
        amt = _num(selling.get("sellingAmount") or line.get("commercialTotal"))
    qty = _int(line.get("qty") or line.get("quantity") or 1)
    uq = _unit_qty(line, qty=qty, rate=None, amount=None)
    if amt is not None and qty and uq:
        try:
            return round(float(amt) / (qty * uq), 2)
        except ZeroDivisionError:
            return None
    return None


def _line_amount(line: Mapping[str, Any]) -> float | None:
    try:
        from WEOS.factory.customer_line_view import customer_line_amount

        return customer_line_amount(line)
    except Exception:
        selling = line.get("selling") if isinstance(line.get("selling"), Mapping) else {}
        return _num(selling.get("sellingAmount") or line.get("commercialTotal") or (line.get("price") or {}).get("total"))


def _type_label(line: Mapping[str, Any]) -> str:
    try:
        from WEOS.factory.customer_line_view import customer_type_label

        return customer_type_label(line) or "Item"
    except Exception:
        return _txt(line.get("displayName") or line.get("product") or "Item") or "Item"


def _type_group(line: Mapping[str, Any]) -> str:
    try:
        from WEOS.factory.line_kind import totals_group_for_line

        return totals_group_for_line(line) or _type_label(line)
    except Exception:
        return _type_label(line)


def _write_letterhead(ws, company: Mapping[str, Any], styles, *, row: int = 1, temps: list[str]) -> int:
    co_name = (_txt(company.get("companyName") or company.get("name") or "WEOS")).upper()
    logo = _logo_png_bytes(company)
    start = row
    col_text = 1
    if logo:
        try:
            _embed_png(ws, logo, f"A{row}", width=88, height=42, temps=temps)
            col_text = 2
            ws.row_dimensions[row].height = 36
        except Exception:
            col_text = 1
    cell = ws.cell(row, col_text, co_name)
    cell.font = styles["title"]
    row += 1
    addr = _txt(company.get("address"))
    if addr:
        ws.cell(row, col_text, addr).font = styles["muted"]
        row += 1
    contact = " · ".join(
        x for x in (_txt(company.get("phone")), _txt(company.get("email")), _txt(company.get("website"))) if x
    )
    if contact:
        ws.cell(row, col_text, contact).font = styles["muted"]
        row += 1
    gst = _txt(company.get("gstNo"))
    if gst:
        ws.cell(row, col_text, f"GSTIN: {gst}").font = styles["muted"]
        row += 1
    if row == start + 1:
        row += 1
    return max(row + 1, start + 3)


def prepare_customer_export_payload(doc: Mapping[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build the same customer-facing payload the quote PDF uses (no factory BOM)."""
    doc = dict(doc or {})
    try:
        from WEOS.factory.project_engine import calculate_project

        result = calculate_project(doc, optimize=False)
    except Exception:
        result = {"lines": list(doc.get("lines") or []), "combined": {}, "price": {}}
    calc_lines = list(result.get("lines") or [])
    orig_lines = list(doc.get("lines") or [])
    by_id = {
        str(ln.get("lineId")): ln
        for ln in orig_lines
        if isinstance(ln, dict) and ln.get("lineId")
    }
    lines: list[dict[str, Any]] = []
    for i, ln in enumerate(calc_lines or orig_lines):
        if not isinstance(ln, dict):
            continue
        src = by_id.get(str(ln.get("lineId") or ""))
        if src is None and i < len(orig_lines) and isinstance(orig_lines[i], dict):
            src = orig_lines[i]
        merged = dict(ln)
        if isinstance(src, dict):
            photo = src.get("designPhoto") if isinstance(src.get("designPhoto"), dict) else None
            if photo and not merged.get("designPhoto"):
                merged["designPhoto"] = dict(photo)
            loc = _txt(src.get("locationName") or src.get("positionName"))
            if not loc and isinstance(src.get("options"), Mapping):
                loc = _txt(src["options"].get("locationName") or src["options"].get("positionName"))
            if loc and not merged.get("locationName"):
                merged["locationName"] = loc
                merged["positionName"] = loc
                opts = dict(merged.get("options") or {}) if isinstance(merged.get("options"), Mapping) else {}
                opts["locationName"] = loc
                opts["positionName"] = loc
                merged["options"] = opts
            for key in ("sellingRate", "saleUnit", "colour", "glass", "displayName", "productType", "serial"):
                if not merged.get(key) and src.get(key) not in (None, ""):
                    merged[key] = src[key]
        lines.append(merged)
    if not lines:
        lines = [dict(ln) for ln in orig_lines if isinstance(ln, dict)]

    gst = _txt(doc.get("companyGst"))
    company: dict[str, Any] = {}
    try:
        from WEOS.factory.company_store import company_branding, load_company, load_company_by_gst, logo_file

        company = dict((load_company_by_gst(gst) if gst else None) or load_company() or {})
        branding = company_branding(gst=gst or None)
        company.update({k: v for k, v in branding.items() if v})
        lf = logo_file()
        if lf:
            company["logoPath"] = str(lf)
    except Exception:
        company = {"companyName": "WEOS"}

    cust_name = _txt(doc.get("customer"))
    profile: dict[str, Any] = {}
    try:
        from WEOS.factory.customer_store import load_customer_profile

        if cust_name:
            profile = dict(load_customer_profile(cust_name) or {})
    except Exception:
        profile = {}
    if doc.get("customerMobile") and not profile.get("phone"):
        profile["phone"] = doc.get("customerMobile")
    if doc.get("customerAddress") and not profile.get("address"):
        profile["address"] = doc.get("customerAddress")
    if doc.get("customerGst") and not profile.get("gstNo"):
        profile["gstNo"] = doc.get("customerGst")

    payload = {
        **result,
        "projectId": doc.get("projectId"),
        "customer": cust_name or _txt(doc.get("customerMobile")) or "—",
        "name": doc.get("name"),
        "quotationId": doc.get("quotationId") or doc.get("quoteNumber") or doc.get("quoteId") or doc.get("projectId"),
        "quoteDate": doc.get("quoteDate") or doc.get("createdAt"),
        "createdOn": doc.get("createdAt"),
        "updatedOn": doc.get("updatedAt"),
        "status": doc.get("status"),
        "description": doc.get("description"),
        "terms": doc.get("terms") or company.get("terms"),
        "companyGst": gst,
        "lines": lines,
        "price": result.get("price") or {},
        "combined": result.get("combined") or {},
        "customerProfile": profile,
        "company": company,
        "sellingIncludesGst": bool(doc.get("sellingIncludesGst")),
    }
    return payload, company


def export_quote_xlsx(
    payload: Mapping[str, Any],
    company: Mapping[str, Any] | None = None,
    *,
    ledger: Mapping[str, Any] | None = None,
    embed_drawings: str = "thumb",
) -> bytes:
    company = dict(company or payload.get("company") or {})
    wb, st = _wb()
    ws = wb.active
    ws.title = "Quote"
    temps: list[str] = []
    try:
        r = _write_letterhead(ws, company, st, temps=temps)
        ws.cell(r, 1, "Customer Quotation").font = st["section"]
        r += 2

        qid = _txt(payload.get("quotationId") or payload.get("quoteNumber") or payload.get("projectId") or "—")
        qdate = _txt(payload.get("quoteDate") or payload.get("createdOn") or payload.get("createdAt"))
        if qdate and "T" in qdate:
            try:
                qdate = datetime.fromisoformat(qdate.replace("Z", "+00:00")).strftime("%d-%m-%Y")
            except Exception:
                qdate = qdate[:10]
        cust = _txt(payload.get("customer"))
        profile = dict(payload.get("customerProfile") or {})
        status = _txt(payload.get("status"))

        meta = [
            ("Quote No", qid),
            ("Date", qdate or datetime.now(timezone.utc).strftime("%d-%m-%Y")),
            ("Customer", cust.upper() if cust and cust != "—" else "—"),
            ("Customer address", _txt(profile.get("address"))),
            (
                "Customer contact",
                " · ".join(x for x in (_txt(profile.get("phone")), _txt(profile.get("email")), _txt(profile.get("contactPerson"))) if x),
            ),
            ("Customer GSTIN", _txt(profile.get("gstNo"))),
            ("Project", _txt(payload.get("name"))),
            ("Status", status.replace("_", " ").title() if status else ""),
        ]
        for label, val in meta:
            if not val:
                continue
            ws.cell(r, 1, f"{label}:").font = st["bold"]
            ws.cell(r, 2, val).alignment = st["wrap"]
            r += 1
        desc = _txt(payload.get("description"))
        if desc:
            ws.cell(r, 1, "Description:").font = st["bold"]
            ws.cell(r, 2, desc).alignment = st["wrap"]
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 36
            r += 1
        r += 1

        headers = ["Serial", "Location", "Type / design", "Specifications", "Qty", "Rate", "Amount", "Drawing"]
        header_row = r
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(r, c, h)
            cell.font = st["head"]
            cell.fill = st["fill_head"]
            cell.alignment = st["wrap"]
        ws.cell(r, 9, "W mm").font = st["muted"]
        ws.cell(r, 10, "H mm").font = st["muted"]
        ws.cell(r, 11, "Unit qty").font = st["muted"]
        ws.cell(r, 12, "Sale unit").font = st["muted"]
        ws.cell(r, 13, "Type group").font = st["muted"]
        r += 1
        first_data = r

        lines = [ln for ln in (payload.get("lines") or []) if isinstance(ln, Mapping)]
        from WEOS.factory.line_kind import design_serial_label, line_location_name

        png_by_i: dict[int, bytes] = {}
        mode = str(embed_drawings or "thumb").strip().lower()
        if mode not in ("0", "none", "off", "false", "skip", "no"):
            try:
                from WEOS.factory.elevation_cache import XLSX_MAX_PX, XLSX_PNG_SCALE, prefetch_line_pngs

                png_by_i = prefetch_line_pngs(
                    list(lines),
                    scale=XLSX_PNG_SCALE if mode != "full" else 1.2,
                    max_px=XLSX_MAX_PX if mode != "full" else 480,
                    max_workers=4,
                )
            except Exception:
                png_by_i = {}

        for i, line in enumerate(lines):
            serial = _txt(line.get("serial") or line.get("serialLabel")) or design_serial_label(i, None)
            if "·" in serial:
                serial = serial.split("·", 1)[0].strip() or f"W{i + 1}"
            loc = line_location_name(line) or "—"
            typ = _type_label(line)
            group = _type_group(line)
            specs = _specs_text(line)
            qty = _int(line.get("qty") or line.get("quantity") or 1)
            rate = _line_rate(line)
            amt = _line_amount(line)
            unit = _sale_unit(line)
            uq = _unit_qty(line, qty=qty, rate=rate, amount=amt)
            try:
                w = float(line.get("width") or 0) or None
                h = float(line.get("height") or 0) or None
            except (TypeError, ValueError):
                w = h = None

            ws.cell(r, 1, serial)
            ws.cell(r, 2, loc).alignment = st["wrap"]
            ws.cell(r, 3, typ).alignment = st["wrap"]
            ws.cell(r, 4, specs).alignment = st["wrap"]
            qcell = ws.cell(r, 5, qty)
            _qty_fmt(qcell)
            rcell = ws.cell(r, 6, round(float(rate), 2) if rate is not None else None)
            _money_fmt(rcell)
            acell = ws.cell(r, 7, f"=IF(OR(E{r}=\"\",F{r}=\"\",K{r}=\"\"),\"\",ROUND(E{r}*K{r}*F{r},2))")
            _money_fmt(acell)
            acell.font = st["bold"]
            ws.cell(r, 9, w)
            ws.cell(r, 10, h)
            kcell = ws.cell(r, 11, _unit_qty_formula_or_value(line, r, uq))
            _unit_fmt(kcell)
            ws.cell(r, 12, unit)
            ws.cell(r, 13, group)
            ws.row_dimensions[r].height = 84
            png = png_by_i.get(i)
            if png:
                try:
                    _embed_png(ws, png, f"H{r}", width=118, height=78, temps=temps)
                except Exception:
                    pass
            r += 1

        last_data = r - 1
        if last_data < first_data:
            ws.cell(r, 1, "No products on this quote").font = st["muted"]
            last_data = first_data
            r += 1

        r += 1
        ws.cell(r, 1, "TOTALS").font = st["section"]
        r += 1
        groups: list[str] = []
        try:
            from WEOS.factory.line_kind import quote_qty_breakdown

            groups = [lab for lab, _qty in quote_qty_breakdown(lines)]
        except Exception:
            groups = []
        extra = []
        for ln in lines:
            g = _type_group(ln)
            if g not in groups and g not in extra:
                extra.append(g)
        for g in groups + extra:
            ws.cell(r, 1, g).font = st["bold"]
            qf = f'=SUMIF($M${first_data}:$M${last_data},A{r},$E${first_data}:$E${last_data})'
            af = f'=SUMIF($M${first_data}:$M${last_data},A{r},$G${first_data}:$G${last_data})'
            qcell = ws.cell(r, 5, qf)
            _qty_fmt(qcell)
            acell = ws.cell(r, 7, af)
            _money_fmt(acell)
            r += 1

        r += 1
        taxable_row = r
        ws.cell(r, 1, "Taxable (ex-GST)").font = st["bold"]
        tcell = ws.cell(r, 7, f"=IF(COUNTA(G{first_data}:G{last_data})=0,0,SUM(G{first_data}:G{last_data}))")
        _money_fmt(tcell)
        r += 1
        gst_pct_row = r
        ws.cell(r, 1, "GST %").font = st["bold"]
        ws.cell(r, 7, _GST_PCT)
        ws.cell(r, 7).number_format = "0.00"
        r += 1
        gst_amt_row = r
        includes = bool(payload.get("sellingIncludesGst"))
        ws.cell(r, 1, f"GST @ {_GST_PCT:g}%").font = st["bold"]
        if includes:
            # Back GST out of taxable (amounts already include GST)
            gcell = ws.cell(
                r,
                7,
                f"=ROUND(G{taxable_row}*G{gst_pct_row}/(100+G{gst_pct_row}),2)",
            )
        else:
            gcell = ws.cell(r, 7, f"=ROUND(G{taxable_row}*G{gst_pct_row}/100,2)")
        _money_fmt(gcell)
        r += 1
        grand_row = r
        ws.cell(r, 1, "Grand total (with GST)").font = st["grand"]
        if includes:
            gtot = ws.cell(r, 7, f"=G{taxable_row}")
        else:
            gtot = ws.cell(r, 7, f"=ROUND(G{taxable_row}+G{gst_amt_row},2)")
        _money_fmt(gtot)
        gtot.font = st["grand"]
        gtot.fill = st["fill_grand"]
        ws.cell(r, 1).fill = st["fill_grand"]
        r += 2

        terms = _txt(payload.get("terms") or company.get("terms"))
        if not terms:
            terms = (
                "1. Specs & sizes may differ 7–9 mm after site measurement.\n"
                "2. Pricing Ex-Works unless noted. GST extra as applicable.\n"
                "3. Payment as agreed. Order confirmation required.\n"
                "4. Delivery typically 3+ weeks from confirmation.\n"
                "5. Quotation valid 15 days.\n"
                "6. Warranty: profile manufacturing defects as per policy."
            )
        ws.cell(r, 1, "Terms & Conditions").font = st["section"]
        r += 1
        ws.cell(r, 1, terms).alignment = st["wrap"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 6, end_column=7)
        ws.row_dimensions[r].height = 96
        last_print_row = r + 6

        widths = {"A": 12, "B": 16, "C": 20, "D": 42, "E": 8, "F": 12, "G": 14, "H": 20, "I": 10, "J": 10, "K": 12, "L": 12, "M": 18}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for col in ("I", "J", "K", "L", "M"):
            ws.column_dimensions[col].hidden = True

        _apply_a4(ws, last_col="H", last_row=last_print_row, title_rows=f"{header_row}:{header_row}")
        _header_footer(ws, company=_txt(company.get("companyName") or "WEOS").upper(), right=f"Quote {qid}")
        ws.freeze_panes = f"A{first_data}"
        ws.sheet_view.showGridLines = False

        try:
            from openpyxl.workbook.defined_name import DefinedName

            for nm, ref in (
                ("QuoteTaxable", f"Quote!$G${taxable_row}"),
                ("QuoteGstPct", f"Quote!$G${gst_pct_row}"),
                ("QuoteGstAmt", f"Quote!$G${gst_amt_row}"),
                ("QuoteGrand", f"Quote!$G${grand_row}"),
            ):
                wb.defined_names.add(DefinedName(name=nm, attr_text=ref))
        except Exception:
            pass

        if ledger:
            _write_ledger_sheet(
                wb,
                ledger,
                company,
                st,
                temps=temps,
                quote_grand_ref=f"Quote!G{grand_row}",
                quote_taxable_ref=f"Quote!G{taxable_row}",
            )
        else:
            _write_ledger_sheet(
                wb,
                {"customer": cust, "advances": [], "projects": [], "totals": {}, "profile": profile},
                company,
                st,
                temps=temps,
                quote_grand_ref=f"Quote!G{grand_row}",
                quote_taxable_ref=f"Quote!G{taxable_row}",
            )

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    finally:
        for p in temps:
            try:
                os.unlink(p)
            except OSError:
                pass


def _write_ledger_sheet(
    wb,
    ledger: Mapping[str, Any],
    company: Mapping[str, Any],
    st,
    *,
    temps: list[str],
    quote_grand_ref: str | None = None,
    quote_taxable_ref: str | None = None,
) -> None:
    ws = None
    try:
        active = wb.active
        if active is not None and active.title in ("Sheet", "Ledger") and active.max_row <= 1:
            ws = active
            ws.title = "Ledger"
    except Exception:
        ws = None
    if ws is None:
        ws = wb.create_sheet("Ledger")
    r = _write_letterhead(ws, company, st, temps=temps)
    ws.cell(r, 1, "Customer account ledger").font = st["section"]
    r += 1
    profile = dict(ledger.get("profile") or {})
    cust = _txt(ledger.get("customer") or profile.get("name"))
    as_of = _txt(ledger.get("asOf"))
    if as_of:
        try:
            as_of = datetime.fromisoformat(as_of.replace("Z", "+00:00")).strftime("%d %b %Y")
        except Exception:
            as_of = as_of[:10]
    ws.cell(r, 1, "Customer:").font = st["bold"]
    ws.cell(r, 2, cust or "—")
    r += 1
    extra = " · ".join(
        x
        for x in (
            _txt(profile.get("address")),
            _txt(profile.get("phone")),
            f"GSTIN {_txt(profile.get('gstNo'))}" if _txt(profile.get("gstNo")) else "",
        )
        if x
    )
    if extra:
        ws.cell(r, 1, extra).font = st["muted"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    if as_of:
        ws.cell(r, 1, f"As of {as_of}").font = st["muted"]
        r += 1
    r += 1

    ws.cell(r, 1, "Running quotes").font = st["section"]
    r += 1
    for c, h in enumerate(["Quote #", "Project", "Version", "Status", "Taxable", "With GST"], start=1):
        cell = ws.cell(r, c, h)
        cell.font = st["head"]
        cell.fill = st["fill_head"]
    r += 1
    proj_first = r
    for p in ledger.get("projects") or []:
        ws.cell(r, 1, _txt(p.get("quotationId") or p.get("projectId")))
        ws.cell(r, 2, _txt(p.get("name")))
        ws.cell(r, 3, p.get("version"))
        ws.cell(r, 4, _txt(p.get("status")))
        tcell = ws.cell(r, 5, _num(p.get("totalTaxable") if p.get("totalTaxable") is not None else p.get("grandTotal")))
        _money_fmt(tcell)
        gcell = ws.cell(r, 6, _num(p.get("totalGrand") if p.get("totalGrand") is not None else p.get("grandTotal")))
        _money_fmt(gcell)
        r += 1
    proj_last = r - 1
    if proj_last < proj_first:
        ws.cell(r, 1, "No running quotes yet").font = st["muted"]
        proj_last = proj_first
        r += 1

    r += 1
    ws.cell(r, 1, "Advances").font = st["section"]
    r += 1
    for c, h in enumerate(["Date", "Amount", "Mode", "Quote", "Version", "Reference", "Note"], start=1):
        cell = ws.cell(r, c, h)
        cell.font = st["head"]
        cell.fill = st["fill_head"]
    r += 1
    adv_first = r
    for a in ledger.get("advances") or []:
        linked = a.get("linkedQuote") or {}
        ws.cell(r, 1, _txt(a.get("paidAt") or a.get("date"))[:10])
        acell = ws.cell(r, 2, _num(a.get("amount")) or 0)
        _money_fmt(acell)
        ws.cell(r, 3, _txt(a.get("paymentMode") or "cash").upper())
        ws.cell(r, 4, _txt(a.get("quoteId") or linked.get("quotationId") or a.get("projectId")))
        ws.cell(r, 5, a.get("quoteVersion") if a.get("quoteVersion") is not None else linked.get("version"))
        ws.cell(r, 6, _txt(a.get("reference")))
        ws.cell(r, 7, _txt(a.get("note")))
        r += 1
    adv_last = r - 1
    if adv_last < adv_first:
        ws.cell(r, 2, 0)
        _money_fmt(ws.cell(r, 2))
        ws.cell(r, 3, "").font = st["muted"]
        adv_last = adv_first
        r += 1

    r += 2
    totals = dict(ledger.get("totals") or {})
    ws.cell(r, 1, "TOTALS").font = st["section"]
    r += 1
    ws.cell(r, 1, "Taxable (quotes)").font = st["bold"]
    if quote_taxable_ref and not (ledger.get("projects") or []):
        tcell = ws.cell(r, 2, f"={quote_taxable_ref}")
    else:
        tcell = ws.cell(r, 2, f"=SUM(E{proj_first}:E{proj_last})")
    _money_fmt(tcell)
    taxable_row = r
    r += 1
    ws.cell(r, 1, "With GST (quotes)").font = st["bold"]
    if quote_grand_ref and not (ledger.get("projects") or []):
        gcell = ws.cell(r, 2, f"={quote_grand_ref}")
    else:
        gcell = ws.cell(r, 2, f"=SUM(F{proj_first}:F{proj_last})")
    _money_fmt(gcell)
    grand_row = r
    r += 1
    ws.cell(r, 1, "Total advances").font = st["bold"]
    acell = ws.cell(r, 2, f"=SUM(B{adv_first}:B{adv_last})")
    _money_fmt(acell)
    adv_row = r
    r += 1
    ws.cell(r, 1, "Balance (taxable − advances)").font = st["bold"]
    bcell = ws.cell(r, 2, f"=ROUND(B{taxable_row}-B{adv_row},2)")
    _money_fmt(bcell)
    r += 1
    ws.cell(r, 1, "Balance (grand − advances)").font = st["grand"]
    bg = ws.cell(r, 2, f"=ROUND(B{grand_row}-B{adv_row},2)")
    _money_fmt(bg)
    bg.font = st["grand"]
    bg.fill = st["fill_grand"]
    ws.cell(r, 1).fill = st["fill_grand"]
    last_row = r
    r += 2
    ws.cell(r, 1, "Taxable is ex-GST; with GST adds 18%. Balance = total − advances.").font = st["muted"]

    for col, w in {"A": 28, "B": 22, "C": 12, "D": 18, "E": 12, "F": 14, "G": 28}.items():
        ws.column_dimensions[col].width = w
    _apply_a4(ws, last_col="G", last_row=last_row + 2)
    _header_footer(
        ws,
        company=_txt(company.get("companyName") or "WEOS").upper(),
        right=f"Ledger · {cust}"[:50],
        footer_left="Powered by WEOS — ledger",
    )
    ws.sheet_view.showGridLines = False
    # Keep stored totals as comments only — formulas are source of truth.
    _ = totals


def export_ledger_xlsx(ledger: Mapping[str, Any], company: Mapping[str, Any] | None = None) -> bytes:
    company = company or {}
    wb, st = _wb()
    temps: list[str] = []
    try:
        _write_ledger_sheet(wb, ledger, company, st, temps=temps)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    finally:
        for p in temps:
            try:
                os.unlink(p)
            except OSError:
                pass


def export_advance_xlsx(
    advance: Mapping[str, Any],
    *,
    company: Mapping[str, Any] | None = None,
    ledger: Mapping[str, Any] | None = None,
    customer: str | None = None,
) -> bytes:
    company = company or {}
    ledger = ledger or {}
    profile = ledger.get("profile") or {}
    totals = ledger.get("totals") or {}
    linked = advance.get("linkedQuote") if isinstance(advance.get("linkedQuote"), Mapping) else {}
    wb, st = _wb()
    ws = wb.active
    ws.title = "Advance Slip"
    temps: list[str] = []
    try:
        r = _write_letterhead(ws, company, st, temps=temps)
        ws.cell(r, 1, "ADVANCE RECEIPT / PAYMENT SLIP").font = st["section"]
        r += 2
        slip_no = _txt(advance.get("slipNo") or advance.get("id") or advance.get("advanceId") or "—")
        paid = _txt(advance.get("paidAt") or advance.get("date"))[:10] or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        cust = _txt(customer or advance.get("customerName") or ledger.get("customer") or profile.get("name"))
        ws.cell(r, 1, "Slip No:").font = st["bold"]
        ws.cell(r, 2, slip_no)
        ws.cell(r, 4, "Date:").font = st["bold"]
        ws.cell(r, 5, paid)
        r += 1
        ws.cell(r, 1, "Received from:").font = st["bold"]
        ws.cell(r, 2, cust.upper() if cust else "—")
        r += 1
        addr = _txt(profile.get("address"))
        if addr:
            ws.cell(r, 2, addr).font = st["muted"]
            r += 1
        contact = " · ".join(x for x in (_txt(profile.get("phone")), _txt(profile.get("email"))) if x)
        if contact:
            ws.cell(r, 2, contact).font = st["muted"]
            r += 1
        if _txt(profile.get("gstNo")):
            ws.cell(r, 2, f"GSTIN: {_txt(profile.get('gstNo'))}").font = st["muted"]
            r += 1
        r += 1
        ws.cell(r, 1, "Project:").font = st["bold"]
        ws.cell(r, 2, _txt(advance.get("projectName") or linked.get("name") or advance.get("projectId")))
        r += 1
        ws.cell(r, 1, "Quote Ref:").font = st["bold"]
        ws.cell(r, 2, _txt(advance.get("quoteId") or linked.get("quotationId") or advance.get("quotationId")))
        r += 2

        ws.cell(r, 1, "Advance amount").font = st["bold"]
        amt_row = r
        acell = ws.cell(r, 2, _num(advance.get("amount")) or 0)
        _money_fmt(acell)
        acell.font = st["grand"]
        acell.fill = st["fill_grand"]
        r += 1
        ws.cell(r, 1, "Payment mode")
        ws.cell(r, 2, _txt(advance.get("paymentMode") or "cash").upper())
        r += 1
        ws.cell(r, 1, "Reference")
        ws.cell(r, 2, _txt(advance.get("reference")))
        r += 1
        ws.cell(r, 1, "Note")
        ws.cell(r, 2, _txt(advance.get("note"))).alignment = st["wrap"]
        r += 2

        ws.cell(r, 1, "Account advances (all)").font = st["section"]
        r += 1
        for c, h in enumerate(["Date", "Amount", "Mode", "Quote / ref"], start=1):
            cell = ws.cell(r, c, h)
            cell.font = st["head"]
            cell.fill = st["fill_head"]
        r += 1
        adv_first = r
        advances = list(ledger.get("advances") or [])
        if not advances:
            advances = [advance]
        for a in advances:
            ws.cell(r, 1, _txt(a.get("paidAt") or a.get("date"))[:10])
            c2 = ws.cell(r, 2, _num(a.get("amount")) or 0)
            _money_fmt(c2)
            ws.cell(r, 3, _txt(a.get("paymentMode") or "cash").upper())
            linked_a = a.get("linkedQuote") if isinstance(a.get("linkedQuote"), Mapping) else {}
            ws.cell(r, 4, _txt(a.get("quoteId") or linked_a.get("quotationId") or a.get("reference") or a.get("note")))
            r += 1
        adv_last = r - 1

        r += 1
        ws.cell(r, 1, "Account summary").font = st["section"]
        r += 1
        ws.cell(r, 1, "Total value (taxable)").font = st["bold"]
        val_row = r
        vcell = ws.cell(r, 2, _num(totals.get("totalTaxable", totals.get("value", totals.get("billed")))) or 0)
        _money_fmt(vcell)
        r += 1
        ws.cell(r, 1, "Total advances").font = st["bold"]
        tot_adv_row = r
        tcell = ws.cell(r, 2, f"=SUM(B{adv_first}:B{adv_last})")
        _money_fmt(tcell)
        r += 1
        ws.cell(r, 1, "Balance outstanding").font = st["grand"]
        bcell = ws.cell(r, 2, f"=ROUND(B{val_row}-B{tot_adv_row},2)")
        _money_fmt(bcell)
        bcell.font = st["grand"]
        bcell.fill = st["fill_grand"]
        last_row = r
        r += 2
        ws.cell(r, 1, "This slip amount (editable):").font = st["muted"]
        ws.cell(r, 2, f"=B{amt_row}")
        _money_fmt(ws.cell(r, 2))

        for col, w in {"A": 28, "B": 18, "C": 14, "D": 28, "E": 14}.items():
            ws.column_dimensions[col].width = w
        _apply_a4(ws, last_col="E", last_row=last_row + 3)
        _header_footer(
            ws,
            company=_txt(company.get("companyName") or "WEOS").upper(),
            right=f"Slip {slip_no}",
            footer_left="Powered by WEOS — advance receipt",
        )
        ws.sheet_view.showGridLines = False
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    finally:
        for p in temps:
            try:
                os.unlink(p)
            except OSError:
                pass


def safe_xlsx_name(*parts: str) -> str:
    bits = []
    for p in parts:
        s = re.sub(r"[^A-Za-z0-9._-]+", "_", _txt(p)).strip("._")
        if s:
            bits.append(s[:40])
    return ("_".join(bits) or "export") + ".xlsx"
