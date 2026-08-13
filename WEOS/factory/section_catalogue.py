"""Windows section catalogue — Excel → series-wise JSON (profile library data layer).

Engineering usage (track / sash / interlock / meeting) is inferred from catalogue
``name`` text and stored on each section so quotation/PDF can annotate like MAR-QT.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Mapping, Sequence

from WEOS.paths import PACKAGE_ROOT, WORKSPACE_ROOT, data_dir, products_dir

HEADER_MARKERS = {
    "section depth mm",
    "width mm",
    "name",
    "standerd length",
    "standard length",
    "weight (kg/mtr)",
    "height limit",
    "design options",
    "wall thikness",
    "wall thickness",
}

# Prefer shorter, stable ids for UI / API
SERIES_ID_ALIASES: dict[str, str] = {
    "25mm-eco-gulf-system-windows-25mm-2-track-3-track": "25mm_eco_gulf",
    "27mm-high-end-domal": "27mm_high_end_domal",
    "29mm-premium-system-windows-euro-grove": "29mm_premium_euro",
    "31-mm-gulf-lux-slim-sreis": "31mm_gulf_lux_slim",
    "32-mm-dual-euro-grove-sreies": "32mm_dual_euro",
    "35mm-gulf-slim-series": "35mm_gulf_slim",
}

# Map catalogue series → existing WEOS product ids (when manufacturing rules exist)
SERIES_PRODUCT_MAP: dict[str, str] = {
    "29mm_premium_euro": "29mm_sliding",
    "35mm_gulf_slim": "35mm_sliding",
    "25mm_eco_gulf": "29mm_sliding",  # nearest active engine until dedicated product
    "27mm_high_end_domal": "29mm_sliding",
    "31mm_gulf_lux_slim": "29mm_sliding",
    "32mm_dual_euro": "29mm_sliding",
}


def sections_dir() -> Path:
    d = products_dir() / "sections"
    d.mkdir(parents=True, exist_ok=True)
    return d


def catalogue_path() -> Path:
    return sections_dir() / "catalogue.json"


def _slug(text: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", (text or "").strip().lower()).strip("-")
    return s or "series"


def _norm_header(cell: Any) -> str:
    return re.sub(r"\s+", " ", str(cell or "").strip().lower())


def _clean_text(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).replace("\u00b0", "°").replace("Degree", "Degree").strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


def _parse_weight_kg_m(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.search(r"([\d.]+)", str(val).replace(",", "."))
    return float(m.group(1)) if m else None


def _parse_height_limit_mm(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.search(r"([\d.]+)", str(val).replace(",", ""))
    return float(m.group(1)) if m else None


def infer_usage(name: str | None) -> str:
    """Map catalogue section name → usage role (frame/track/sash/interlock/…)."""
    n = (name or "").lower()
    if not n:
        return "unknown"
    if "meeting" in n:
        return "meeting"
    if "interlock" in n or "renf" in n or re.search(r"\bint\.?\b", n):
        return "interlock"
    if "shutter" in n or "sash" in n:
        return "sash"
    if "domal" in n:
        return "track"
    if "track" in n or "trck" in n:
        if any(k in n for k in ("horiz", "top", "bottom", "tp ")):
            return "track_horizontal"
        if any(k in n for k in ("vert", "left", "right")):
            return "track_vertical"
        if "add" in n or "4" in n:
            return "track_add"
        return "track"
    if "frame" in n or "outer" in n:
        return "frame"
    return "other"


def parse_track_count(name: str | None) -> float | None:
    """Extract the track count from a section name (``2 track`` → 2.0,
    ``2.5 track`` → 2.5, ``3 track`` → 3.0). Returns None when not a track row."""
    n = (name or "").lower()
    if "track" not in n and "trck" not in n:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:track|trck)", n)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def parse_glass_options(name: str | None) -> list[str]:
    """Map glass codes in a section name to availability.

    ``sg`` → single glass; ``dg`` / ``gd`` → DGU (double-glazed unit). A section
    tagged ``sg, dg`` therefore supports both single and DGU glass.
    """
    n = (name or "").lower()
    opts: list[str] = []
    # token match so we don't catch letters inside words
    tokens = set(re.findall(r"[a-z]+", n))
    if "sg" in tokens:
        opts.append("single")
    if "dg" in tokens or "gd" in tokens or "dgu" in tokens:
        opts.append("dgu")
    return opts


_SG_DG_PRINT = re.compile(r"(?i)(?:^|[\s,;/]+)(?:sg|dg|gd|dgu)\b")
_TRACK_OPTION_DUMP = re.compile(
    r"(?i)(?:[,;/\-–]\s*)?(?:\d+(?:\.\d+)?\s*[-\s]*track(?:s)?)"
    r"(?:\s*[,;/&]\s*(?:\d+(?:\.\d+)?\s*[-\s]*track(?:s)?))+"
)
_SINGLE_TRACK_TOKEN = re.compile(r"(?i)\b\d+(?:\.\d+)?\s*[-\s]*track(?:s)?\b")
_TRACK_ORIENT_FLUFF = re.compile(
    r"(?i)\b(?:all\s+side|top/?bottom|left/?right|horizont[ae]l?e?|verticale?|high\s*end|domal)\b"
)


def has_track_option_dump(text: Any) -> bool:
    """True when a string lists more than one track option (e.g. ``2 track, 3 track``)."""
    s = str(text or "")
    if _TRACK_OPTION_DUMP.search(s):
        return True
    return len(_SINGLE_TRACK_TOKEN.findall(s)) >= 2


def clean_series_print_name(name: str | None) -> str:
    """Strip catalogue track-option dumps from a series / product title.

    May return empty when the name was only track tokens (e.g. ``3 track``).
    """
    raw = str(name or "").strip()
    if not raw:
        return ""
    cleaned = _TRACK_OPTION_DUMP.sub(" ", raw)
    cleaned = _SINGLE_TRACK_TOKEN.sub(" ", cleaned)
    return re.sub(r"[\s,;]+", " ", cleaned).strip(" ,;/-")


def format_active_track_print(
    track_count: Any,
    track_sec: Mapping[str, Any] | None = None,
    *,
    wall_mm: Any = None,
) -> str:
    """Customer TRACK line: only the selected count, never ``2 track, 3 track``."""
    tc_lbl = ""
    try:
        if track_count is not None and str(track_count).strip() != "":
            tc_lbl = f"{float(track_count):g}-track"
    except (TypeError, ValueError):
        tc_lbl = ""
    dim = None
    extra_name = ""
    if isinstance(track_sec, Mapping):
        d, w = track_sec.get("sectionDepthMm"), track_sec.get("widthMm")
        if d is not None and w is not None:
            dim = f"{float(d):g}×{float(w):g} mm"
        extra_name = clean_series_print_name(track_sec.get("name") or "")
        extra_name = _TRACK_ORIENT_FLUFF.sub(" ", extra_name)
        extra_name = re.sub(r"[\s,;]+", " ", extra_name).strip(" ,;/-")
        if extra_name.lower() in ("track", "frame", "outer", ""):
            extra_name = ""
        if wall_mm is None:
            wall_mm = track_sec.get("wallThicknessMm")
    bits: list[str] = []
    if tc_lbl:
        bits.append(tc_lbl)
    if extra_name:
        bits.append(extra_name)
    if dim:
        bits.append(dim)
    if wall_mm not in (None, ""):
        wtxt = f"{wall_mm:g}" if isinstance(wall_mm, (int, float)) else str(wall_mm)
        bits.append(f"wall {wtxt} mm")
    return " · ".join(bits)


def clean_profile_print_name(name: str | None) -> str:
    """Customer print name without dual ``sg, dg`` tags."""
    raw = str(name or "").strip()
    if not raw:
        return ""
    cleaned = _SG_DG_PRINT.sub(" ", raw)
    cleaned = re.sub(r"[\s,;]+", " ", cleaned).strip(" ,;/-")
    return cleaned or raw


def section_glass_family_rank(sec: Mapping[str, Any] | None, family: str | None) -> int:
    """Higher is better. -1 = incompatible with requested SG/DG family."""
    if not isinstance(sec, Mapping):
        return -1
    opts = list(sec.get("glassOptions") or []) or parse_glass_options(sec.get("name"))
    fam = str(family or "").strip().lower()
    if fam in ("sg", "single", "laminated", "lami"):
        fam = "single"
    elif fam in ("dg", "dgu", "igu", "gd", "double"):
        fam = "dgu"
    if not fam:
        return 0
    if not opts:
        return 1
    if fam in opts and len(opts) == 1:
        return 3
    if fam in opts:
        return 2
    return -1


def _print_dim(sec: Mapping[str, Any] | None, *, clean_names: bool = True) -> str | None:
    if not isinstance(sec, Mapping):
        return None
    d, w = sec.get("sectionDepthMm"), sec.get("widthMm")
    if d is None or w is None:
        return None
    name = clean_profile_print_name(sec.get("name")) if clean_names else str(sec.get("name") or "")
    wall = sec.get("wallThicknessMm")
    bits = []
    if name:
        bits.append(name)
    bits.append(f"{d:g}×{w:g} mm")
    if wall not in (None, ""):
        bits.append(f"wall {wall:g} mm")
    return " · ".join(bits)


def is_center_opening_only(name: str | None) -> bool:
    n = (name or "").lower()
    return "center" in n and ("open" in n or "opning" in n or "opening" in n)


def _usage_label(usage: str) -> str:
    return {
        "track": "Track / Outer frame",
        "track_horizontal": "Track — horizontal",
        "track_vertical": "Track — vertical",
        "track_add": "Add track",
        "sash": "Sash / Shutter",
        "interlock": "Interlock",
        "meeting": "Meeting section",
        "frame": "Frame",
        "other": "Section",
        "unknown": "Section",
    }.get(usage, usage)


def _is_header_row(row: Sequence[Any]) -> bool:
    cells = {_norm_header(c) for c in row if c is not None and str(c).strip()}
    return "section depth mm" in cells and "width mm" in cells


def _is_series_title_row(row: Sequence[Any]) -> bool:
    vals = [c for c in row if c is not None and str(c).strip() != ""]
    if len(vals) != 1:
        return False
    text = _norm_header(vals[0])
    if text in HEADER_MARKERS or "section depth" in text:
        return False
    # Title-like: contains mm / series / system / track
    return bool(re.search(r"\d+\s*mm|series|system|track|domal|gulf|euro", text, re.I))


def _is_section_row(row: Sequence[Any]) -> bool:
    depth = row[0] if len(row) > 0 else None
    width = row[1] if len(row) > 1 else None
    try:
        float(depth)
        float(width)
        return True
    except (TypeError, ValueError):
        return False


def find_excel_path(explicit: str | Path | None = None) -> Path:
    if explicit:
        p = Path(explicit)
        if p.is_file():
            return p
        raise FileNotFoundError(f"Excel not found: {explicit}")
    candidates = [
        WORKSPACE_ROOT / "deta windows.xlsx",
        WORKSPACE_ROOT / "references" / "deta windows.xlsx",
        Path(r"d:\Downloads\deta windows.xlsx"),
        PACKAGE_ROOT.parent / "deta windows.xlsx",
    ]
    for c in candidates:
        if c.is_file():
            return c
    raise FileNotFoundError(
        "deta windows.xlsx not found. Place it in workspace root or references/."
    )


def parse_excel(path: str | Path | None = None) -> dict[str, Any]:
    """Parse DETA windows catalogue Excel into series-wise structure."""
    from openpyxl import load_workbook

    xlsx = find_excel_path(path)
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    series_list: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    headers_seen = False

    for row in ws.iter_rows(values_only=True):
        cells = list(row[:8]) if row else []
        if not any(c is not None and str(c).strip() != "" for c in cells):
            continue

        if _is_series_title_row(cells):
            title = _clean_text(cells[0]) or "Untitled series"
            raw_id = _slug(title)
            sid = SERIES_ID_ALIASES.get(raw_id, raw_id.replace("-", "_"))
            # Prefer short aliases even when slug differs slightly
            for alias_key, alias_id in SERIES_ID_ALIASES.items():
                if alias_key in raw_id or raw_id in alias_key:
                    sid = alias_id
                    break
            # Manual fixes for known titles
            tl = title.lower()
            if "25mm" in tl and "eco" in tl:
                sid = "25mm_eco_gulf"
            elif "27mm" in tl or ("high end" in tl and "domal" in tl):
                sid = "27mm_high_end_domal"
            elif "29mm" in tl:
                sid = "29mm_premium_euro"
            elif "31" in tl and "gulf" in tl:
                sid = "31mm_gulf_lux_slim"
            elif "32" in tl:
                sid = "32mm_dual_euro"
            elif "35mm" in tl:
                sid = "35mm_gulf_slim"

            current = {
                "id": sid,
                "title": title,
                "displayName": title,
                "productId": SERIES_PRODUCT_MAP.get(sid),
                "sections": [],
                "designOptions": [],
            }
            series_list.append(current)
            headers_seen = False
            continue

        if _is_header_row(cells):
            headers_seen = True
            continue

        if current is None:
            continue

        if _is_section_row(cells):
            name = _clean_text(cells[2]) if len(cells) > 2 else None
            # Skip nameless incomplete rows (e.g. trailing depth/width only)
            if not name and len(series_list[-1]["sections"]) >= 4:
                # still record if both dims present and earlier sections exist — keep if name missing only at end
                if not name:
                    continue
            usage = infer_usage(name)
            weight_raw = cells[4] if len(cells) > 4 else None
            design_opt = _clean_text(cells[6]) if len(cells) > 6 else None
            section = {
                "id": f"{current['id']}__{_slug(name or 'section')}",
                "name": name or "Unnamed",
                "usage": usage,
                "usageLabel": _usage_label(usage),
                "sectionDepthMm": float(cells[0]),
                "widthMm": float(cells[1]),
                "standardLength": _clean_text(cells[3]) if len(cells) > 3 else None,
                "weightKgPerMtrRaw": _clean_text(weight_raw),
                "weightKgPerMtr": _parse_weight_kg_m(weight_raw),
                "heightLimit": _clean_text(cells[5]) if len(cells) > 5 else None,
                "heightLimitMm": _parse_height_limit_mm(cells[5] if len(cells) > 5 else None),
                "wallThicknessMm": (
                    float(cells[7])
                    if len(cells) > 7 and cells[7] is not None and str(cells[7]).strip() != ""
                    else None
                ),
                "designNote": design_opt,
                # Excel data model: track count (2 / 2.5 / 3), glass availability
                # (sg=single, dg/gd=DGU), and center-opening-only meeting sections.
                "trackCount": parse_track_count(name),
                "glassOptions": parse_glass_options(name),
                "centerOpeningOnly": is_center_opening_only(name),
            }
            current["sections"].append(section)
            if design_opt and design_opt not in current["designOptions"]:
                current["designOptions"].append(design_opt)
            continue

        # Design-option / note rows (often only column G populated)
        note = None
        for c in cells:
            t = _clean_text(c)
            if t and _norm_header(t) not in HEADER_MARKERS:
                note = t
                break
        if note and current is not None and note not in current["designOptions"]:
            current["designOptions"].append(note)

    wb.close()

    # Drop empty series
    series_list = [s for s in series_list if s["sections"]]

    return {
        "source": xlsx.name,
        "sourcePath": str(xlsx),
        "version": 1,
        "seriesCount": len(series_list),
        "sectionCount": sum(len(s["sections"]) for s in series_list),
        "series": series_list,
    }


def persist_catalogue(doc: Mapping[str, Any]) -> Path:
    """Write catalogue.json + one JSON file per series under products/sections/."""
    root = sections_dir()
    path = catalogue_path()
    path.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    for series in doc.get("series") or []:
        sid = series["id"]
        series_path = root / f"{sid}.json"
        series_path.write_text(
            json.dumps(series, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )

    # Writable mirror under data_dir for runtime overrides
    mirror = data_dir() / "sections"
    mirror.mkdir(parents=True, exist_ok=True)
    (mirror / "catalogue.json").write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
    return path


def _library_catalogue_block(series: Mapping[str, Any]) -> dict[str, Any]:
    """Build an editable catalogue block for a Product Library entry from a series."""
    sections = list(series.get("sections") or [])
    profiles = []
    for sec in sections:
        name = sec.get("name")
        tc = sec.get("trackCount")
        if tc is None:
            tc = parse_track_count(name)
        gopts = list(sec.get("glassOptions") or []) or parse_glass_options(name)
        profiles.append(
            {
                "name": name,
                "usage": sec.get("usage"),
                "usageLabel": sec.get("usageLabel"),
                "sectionDepthMm": sec.get("sectionDepthMm"),
                "widthMm": sec.get("widthMm"),
                "weightKgPerMtr": sec.get("weightKgPerMtr"),
                "standardLength": sec.get("standardLength"),
                "wallThicknessMm": sec.get("wallThicknessMm"),
                "trackCount": tc,
                "glassOptions": gopts,
                "centerOpeningOnly": bool(
                    sec.get("centerOpeningOnly")
                    if sec.get("centerOpeningOnly") is not None
                    else is_center_opening_only(name)
                ),
            }
        )
    # Representative standard length (first non-empty)
    std_len = next((s.get("standardLength") for s in sections if s.get("standardLength")), None)
    depth = next((s.get("sectionDepthMm") for s in sections if s.get("sectionDepthMm")), None)

    # ── Excel data model ────────────────────────────────────────────────────
    # Track is an OPTION: 2 / 2.5 / 3 track are variants of the SAME series that
    # only change the outer/track section; the sash/shutter frame is shared.
    # Stale section JSON (pre-trackCount field) still works via name parse.
    tracks: list[dict[str, Any]] = []
    seen_tc: set[float] = set()
    for sec in sections:
        usage = str(sec.get("usage") or "")
        if usage and usage not in ("track", "track_horizontal", "track_vertical", "frame"):
            continue
        tc = sec.get("trackCount")
        if tc is None:
            tc = parse_track_count(sec.get("name"))
        if tc is None:
            continue
        try:
            tc_f = float(tc)
        except (TypeError, ValueError):
            continue
        if tc_f in seen_tc:
            continue
        seen_tc.add(tc_f)
        tracks.append(
            {
                "count": tc_f,
                "name": sec.get("name"),
                "sectionDepthMm": sec.get("sectionDepthMm"),
                "widthMm": sec.get("widthMm"),
                "sectionId": sec.get("id"),
                # mesh needs 2.5 or 3 track (single 2-track has no room for a mesh sash)
                "meshCapable": tc_f >= 2.5,
            }
        )
    tracks.sort(key=lambda t: t["count"])
    track_counts = [t["count"] for t in tracks]

    # Shared sash / shutter frame (same across all tracks).
    shared_sash = next((s for s in sections if s.get("usage") == "sash"), None)

    # Glass types available anywhere in the series (sg=single, dg/gd=DGU).
    glass_types: list[str] = []
    for sec in sections:
        gopts = list(sec.get("glassOptions") or []) or parse_glass_options(sec.get("name"))
        for g in gopts:
            if g not in glass_types:
                glass_types.append(g)

    mesh_min_track = next((t["count"] for t in tracks if t.get("meshCapable")), None)
    default_track = track_counts[0] if track_counts else None

    return {
        "sectionSeries": series.get("id"),
        "saleUnit": "sqft",
        "ratePerUnit": 0,
        "sizeMm": depth,
        "standardLength": std_len,
        "weightKgPerMtr": next((s.get("weightKgPerMtr") for s in sections if s.get("weightKgPerMtr")), None),
        "profiles": profiles,
        "designOptions": list(series.get("designOptions") or []),
        # Editable Excel data model → round-trips through the product store.
        "tracks": tracks,
        "trackOptions": track_counts,
        "trackCount": default_track,
        "sharedSash": (
            {
                "name": shared_sash.get("name"),
                "widthMm": shared_sash.get("widthMm"),
                "sectionDepthMm": shared_sash.get("sectionDepthMm"),
                "glassOptions": list(shared_sash.get("glassOptions") or [])
                or parse_glass_options(shared_sash.get("name")),
            }
            if shared_sash
            else None
        ),
        "glassTypes": glass_types,
        "compatibility": {
            "trackOptions": track_counts,
            "meshMinTrack": mesh_min_track,
            "glassTypes": glass_types,
            "notes": [
                n
                for n in [
                    ("Mesh requires 2.5 or 3 track" if mesh_min_track else None),
                    ("sg = single glass; dg/gd = DGU (double-glazed)" if glass_types else None),
                ]
                if n
            ],
        },
        "source": "excel_section_catalogue",
        "productId": series.get("productId") or SERIES_PRODUCT_MAP.get(str(series.get("id") or "")),
    }


def sync_catalogue_to_library() -> dict[str, Any]:
    """Feed imported section series INTO the single Product Library (idempotent).

    Every series becomes an editable Product Library entry so it is selectable in
    the Window Cart and manageable in Admin · Product Library. Existing products
    (with manufacturing rules) are never overwritten — only their catalogue block
    and section specs are refreshed.
    """
    import json as _json

    from WEOS.factory.section_catalogue import specs_summary_for_series

    doc = load_catalogue()
    created: list[str] = []
    updated: list[str] = []
    for series in doc.get("series") or []:
        sid = series.get("id")
        if not sid:
            continue
        pdir = products_dir() / sid
        meta_path = pdir / "product.json"
        catalogue_block = _library_catalogue_block(series)
        try:
            specs = specs_summary_for_series(sid)
        except Exception:
            specs = {}
        _tracks = catalogue_block.get("trackOptions") or []
        _glass = catalogue_block.get("glassTypes") or []
        spec_fields = {
            "profileSeries": series.get("title"),
            "sectionSizeMm": catalogue_block.get("sizeMm"),
            "standardLength": catalogue_block.get("standardLength"),
            "wallThicknessMm": specs.get("wallThicknessMm"),
            "track": specs.get("track"),
            "sash": specs.get("sash"),
            "interlock": specs.get("interlock"),
            "trackOptions": (", ".join(f"{t:g} track" for t in _tracks) if _tracks else None),
            "glassTypes": (
                ", ".join({"single": "Single glass", "dgu": "DGU"}.get(g, g) for g in _glass)
                if _glass
                else None
            ),
        }
        spec_fields = {k: v for k, v in spec_fields.items() if v is not None}

        if meta_path.is_file():
            try:
                meta = _json.loads(meta_path.read_text(encoding="utf-8-sig"))
            except Exception:
                meta = {"id": sid}
            meta["catalogue"] = catalogue_block
            meta["sectionSeries"] = sid
            meta.setdefault("linkedProductId", SERIES_PRODUCT_MAP.get(sid))
            # Catalogue series are sliding windows — never leave productType as the
            # opaque "section_series" (that blocked system inference → live preview).
            pt = str(meta.get("productType") or "").strip().lower()
            if not pt or pt in ("section_series", "unknown", "other"):
                meta["productType"] = "sliding"
            if not meta.get("category"):
                meta["category"] = "Windows"
            specs_existing = dict(meta.get("specifications") or {})
            for k, v in spec_fields.items():
                # Refresh derived track/glass options; keep other manual overrides.
                if k in ("trackOptions", "glassTypes", "track", "sash", "interlock") or k not in specs_existing:
                    specs_existing[k] = v
            meta["specifications"] = specs_existing
            meta_path.write_text(_json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
            updated.append(sid)
            continue

        pdir.mkdir(parents=True, exist_ok=True)
        (pdir / "rules").mkdir(exist_ok=True)
        meta = {
            "id": sid,
            "displayName": series.get("displayName") or series.get("title") or sid,
            "productType": "sliding",
            "category": "Windows",
            "units": "mm",
            "version": 1,
            "status": "stub",
            "description": f"Imported from section catalogue — {series.get('title') or sid}.",
            "tagline": series.get("title") or "",
            "warranty": "",
            "heroImage": "/static/products/placeholder.svg",
            "gallery": [],
            "sectionDrawings": [],
            "specifications": spec_fields,
            "materials": [],
            "formulas": {},
            "pdfLayout": {"customer": "marqt_customer", "factory": "woodenmax_factory"},
            "brand": "woodenmax",
            "sectionSeries": sid,
            "linkedProductId": SERIES_PRODUCT_MAP.get(sid) or "29mm_sliding",
            "catalogue": catalogue_block,
        }
        meta_path.write_text(_json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        (pdir / "rules" / "quotation.json").write_text(
            _json.dumps(
                {
                    "currency": "INR",
                    "labourPerOpening": 0,
                    "markupPercent": 15,
                    "gstPercent": 18,
                    "stub": True,
                    "manualRatePerOpening": 0,
                    "rates": {},
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        created.append(sid)
    return {"ok": True, "created": created, "updated": updated, "count": len(created) + len(updated)}


def import_excel(path: str | Path | None = None, *, sync_library: bool = True) -> dict[str, Any]:
    doc = parse_excel(path)
    persist_catalogue(doc)
    result = {
        "ok": True,
        "path": str(catalogue_path()),
        "seriesCount": doc["seriesCount"],
        "sectionCount": doc["sectionCount"],
        "series": [{"id": s["id"], "title": s["title"], "sections": len(s["sections"])} for s in doc["series"]],
    }
    if sync_library:
        try:
            result["library"] = sync_catalogue_to_library()
        except Exception as exc:  # pragma: no cover - defensive
            result["library"] = {"ok": False, "error": str(exc)}
    # Persist imported catalogue + product folders to the durable DB store so the
    # Excel import survives Railway redeploys (ephemeral filesystem).
    try:
        from WEOS.db.product_store import snapshot_all, snapshot_dir

        snapshot_dir("sections")
        snapshot_all()
        result["persisted"] = True
    except Exception as exc:  # pragma: no cover - best-effort
        result["persisted"] = False
        result["persistError"] = str(exc)
    return result


def load_catalogue() -> dict[str, Any]:
    path = catalogue_path()
    if not path.is_file():
        # Auto-import from known Excel locations
        try:
            import_excel()
        except FileNotFoundError:
            return {"source": None, "series": [], "seriesCount": 0, "sectionCount": 0}
    return json.loads(path.read_text(encoding="utf-8-sig"))


def list_series() -> list[dict[str, Any]]:
    doc = load_catalogue()
    out = []
    for s in doc.get("series") or []:
        out.append(
            {
                "id": s["id"],
                "title": s.get("title"),
                "displayName": s.get("displayName") or s.get("title"),
                "productId": s.get("productId") or SERIES_PRODUCT_MAP.get(s["id"]),
                "sectionCount": len(s.get("sections") or []),
                "usages": sorted({sec.get("usage") for sec in (s.get("sections") or [])}),
            }
        )
    return out


def get_series(series_id: str) -> dict[str, Any]:
    path = sections_dir() / f"{series_id}.json"
    if path.is_file():
        doc = json.loads(path.read_text(encoding="utf-8-sig"))
    else:
        doc = None
        cat = load_catalogue()
        for s in cat.get("series") or []:
            if s.get("id") == series_id:
                doc = s
                break
        if doc is None:
            raise FileNotFoundError(f"Section series '{series_id}' not found")
    # Backfill track/glass fields on stale series JSON (pre-Excel-model imports).
    for sec in doc.get("sections") or []:
        if not isinstance(sec, dict):
            continue
        if sec.get("trackCount") is None:
            tc = parse_track_count(sec.get("name"))
            if tc is not None:
                sec["trackCount"] = tc
        if not sec.get("glassOptions"):
            gopts = parse_glass_options(sec.get("name"))
            if gopts:
                sec["glassOptions"] = gopts
        if "centerOpeningOnly" not in sec:
            sec["centerOpeningOnly"] = is_center_opening_only(sec.get("name"))
    return doc


def sections_for_usage(series_id: str, usage: str | None = None) -> list[dict[str, Any]]:
    series = get_series(series_id)
    secs = list(series.get("sections") or [])
    if usage:
        secs = [s for s in secs if s.get("usage") == usage]
    return secs


def specs_summary_for_series(
    series_id: str | None,
    *,
    glass_family: str | None = None,
    track_count: float | None = None,
    clean_names: bool = True,
) -> dict[str, Any]:
    """Build specification fields from catalogue sections.

    ``glass_family``: ``single`` (SG, incl. laminated) or ``dgu`` (DG/IGU).
    When set, pick the matching SG or DG profile — never a dual ``sg, dg`` dump.
    """
    if not series_id:
        return {}
    try:
        series = get_series(series_id)
    except FileNotFoundError:
        return {}

    by_usage: dict[str, list[dict[str, Any]]] = {}
    for sec in series.get("sections") or []:
        by_usage.setdefault(sec.get("usage") or "other", []).append(sec)

    fam = str(glass_family or "").strip().lower()
    if fam in ("sg", "laminated", "lami"):
        fam = "single"
    elif fam in ("dg", "igu", "gd", "double"):
        fam = "dgu"

    def pick(usages: Sequence[str]) -> dict[str, Any] | None:
        ranked: list[tuple[int, dict[str, Any]]] = []
        for u in usages:
            for sec in by_usage.get(u) or []:
                if not isinstance(sec, Mapping):
                    continue
                # Prefer the track count the cart actually uses.
                if track_count is not None and u.startswith("track"):
                    stc = sec.get("trackCount")
                    if stc is None:
                        stc = parse_track_count(sec.get("name"))
                    try:
                        if stc is not None and abs(float(stc) - float(track_count)) > 0.05:
                            continue
                    except (TypeError, ValueError):
                        pass
                rank = section_glass_family_rank(sec, fam or None)
                if rank < 0:
                    continue
                ranked.append((rank, dict(sec)))
        if not ranked:
            for u in usages:
                if by_usage.get(u):
                    return dict(by_usage[u][0])
            return None
        ranked.sort(key=lambda x: -x[0])
        return ranked[0][1]

    track = pick(("track", "track_horizontal", "track_vertical", "frame"))
    frame = pick(("frame", "track"))
    sash = pick(("sash",))
    interlock = pick(("interlock",))
    meeting = pick(("meeting",))

    wall_track = (track or {}).get("wallThicknessMm") if track else None
    wall_sash = (sash or {}).get("wallThicknessMm") if sash else None
    wall_frame = (frame or {}).get("wallThicknessMm") if frame else None

    tc_used = track_count
    if tc_used is None and isinstance(track, Mapping):
        tc_used = track.get("trackCount")
        if tc_used is None:
            tc_used = parse_track_count(track.get("name"))
    active_track = format_active_track_print(tc_used, track, wall_mm=wall_track)
    series_title = clean_series_print_name(series.get("title")) if has_track_option_dump(series.get("title")) else series.get("title")
    if not series_title:
        series_title = series.get("title")

    return {
        "seriesId": series["id"],
        "seriesTitle": series_title,
        "glassFamily": fam or None,
        "track": active_track or _print_dim(track, clean_names=clean_names),
        "trackPrint": active_track or _print_dim(track, clean_names=True),
        "trackCount": float(tc_used) if tc_used is not None else None,
        "frame": _print_dim(frame, clean_names=clean_names),
        "framePrint": _print_dim(frame, clean_names=True),
        "sash": _print_dim(sash, clean_names=clean_names),
        "sashPrint": _print_dim(sash, clean_names=True),
        "interlock": _print_dim(interlock, clean_names=clean_names),
        "interlockPrint": _print_dim(interlock, clean_names=True),
        "meeting": _print_dim(meeting, clean_names=clean_names),
        "wallThicknessMm": wall_track or wall_frame or wall_sash
        or next(
            (s.get("wallThicknessMm") for s in (series.get("sections") or []) if s.get("wallThicknessMm")),
            None,
        ),
        "trackWallMm": wall_track,
        "sashWallMm": wall_sash,
        "frameWallMm": wall_frame,
        "sections": series.get("sections") or [],
        "designOptions": series.get("designOptions") or [],
    }


def ensure_catalogue_imported() -> dict[str, Any]:
    """Idempotent: import Excel if catalogue missing or empty."""
    path = catalogue_path()
    if path.is_file():
        doc = json.loads(path.read_text(encoding="utf-8-sig"))
        if doc.get("series"):
            return {"ok": True, "imported": False, "seriesCount": doc.get("seriesCount", 0)}
    return {**import_excel(), "imported": True}
