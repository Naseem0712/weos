"""Smoke: multi-stage Excel import + optional advance slip (no auto-open)."""
from __future__ import annotations

import os
import sys
import tempfile
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from WEOS.factory.package_quote import normalize_package_quote
from WEOS.factory.project_import import parse_excel_bytes, parse_upload


def _ok(cond: bool, msg: str) -> None:
    if not cond:
        raise SystemExit(f"FAIL: {msg}")
    print("OK:", msg)


def _tiny_xlsx() -> bytes:
    wb = Workbook()
    acc = wb.active
    acc.title = "Account_payment - Demo"
    acc["A1"] = "Demo Accounts"
    acc.append([])
    acc["A2"] = "Name"
    acc["B2"] = "Amount"
    acc["C2"] = "Firmat"
    acc["D2"] = "Date"
    acc.append([1, "Naseem", 105000, "Cash"])
    acc.append([2, "Jahageer", 100000, "Cash", "2024-09-25"])
    acc.append(["Total", 205000])
    acc.append(["Value", 236000])
    acc.append(["Balance", 31000])

    p1 = wb.create_sheet("29MM luxury - 35 mm Luxury Wind")
    p1["A1"] = "CLINT / COMPANY :- Mr Demo ji    TITEL :- 29 mm series Luxury Windows . P. I Number :- 24/25AB291"
    p1["A2"] = "Page 1"
    p1.append([])
    p1.append(["Description", "Width", "Height", "Count", "Area", "Rate", "Amount"])
    p1.append(["GF SR", "Sliding window 8MM", 5, 5, 2, 50, 920, 46000])
    p1.append(["Vent", "Vent 50MM bathroom", 2, 2, 1, 4, 620, 2480])

    p2 = wb.create_sheet("29MM Profiles  - 35 mm Luxury W")
    p2["A1"] = "Page 2"
    p2.append(["Products", "Width", "Height", "Count", "Area", "Rate", "Amount"])
    p2.append(["CH1", "Sliding window page 2", 4, 8, 1, 32, 950, 30400])
    p2.append([3, "Total", 78880])
    p2.append(["Payment Term ;", "GST", 0.18, 14198.4])
    p2.append(["x", "G. Total", 93078.4])

    grill = wb.create_sheet("Grills - 35 mm Luxury Windows")
    grill.append(["Products", "Width", "Height", "Count", " Area  SFT", "Rate", "Amount"])
    grill.append(["GF", "Aluminium Safety Grills", 1.54, 1.44, 1, 23.87, 305, 7280])
    grill.append([2, 1.15, 1.39, 1, 17.2, 305, 5246])
    grill.append([3, "Total", 12526])
    grill.append(["GST", 0.18, 2254.68])
    grill.append(["G. Total", 14780.68])

    bom = wb.create_sheet("Sheet 1 - Rajkumar")
    bom.append(["Rajkumar"])
    bom.append(["Amount"])
    bom.append([1, "Outer big v-5101", 7, 34.14, 1.727, 58.97])

    glass = wb.create_sheet("Glass size - WoodenMax-Project")
    glass.append(["W", "H"])
    glass.append([1200, 1500])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> None:
    raw = _tiny_xlsx()
    pack = parse_excel_bytes(raw, "demo.xlsx")
    _ok(pack["advanceCount"] == 2, f"2 advances got {pack['advanceCount']} {pack.get('advances')}")
    _ok(abs(pack["advanceTotal"] - 205000) < 0.05, f"advance total 205000 got {pack['advanceTotal']}")
    titles = " | ".join(str(s.get("title") or "") for s in pack.get("stages") or [])
    _ok(pack["quoteCount"] == 2, f"windows+grills merged to 2 quotes got {pack['quoteCount']} :: {titles}")
    win = pack["quotes"][0]
    notes = " ".join(str(it.get("note") or "") for it in win.get("items") or [])
    _ok("page 2" in notes.lower() or "Sliding window page 2" in notes, f"page2 items merged: {notes}")
    _ok(pack["advances"][0].get("paidAt") not in {"1", "2"}, f"serial not used as date {pack['advances']}")
    _ok(len(win.get("items") or []) >= 3, f"merged windows items >=3 got {len(win.get('items') or [])}")
    _ok(abs(float(win.get("gstPercent") or 0) - 18) < 0.1, f"windows GST 18 got {win.get('gstPercent')}")
    grill_q = pack["quotes"][1]
    _ok(len(grill_q.get("items") or []) >= 2, f"grill continuation row kept, items={len(grill_q.get('items') or [])}")
    _ok(pack["projectValue"] > 100000, f"project value got {pack['projectValue']}")

    n = normalize_package_quote(
        {
            "gstMode": "exclude",
            "gstPercent": 0,
            "gstAmount": 31090,
            "projectValue": 408377.5,
            "items": [{"category": "other", "amount": 377287.5, "unit": "sft", "note": "ACP"}],
        }
    )
    _ok(n and abs(n["gstAmount"] - 31090) < 0.05, f"imported GST bill kept {n and n.get('gstAmount')}")
    _ok(n and abs(n["projectValue"] - 408377.5) < 0.05, f"imported grand kept {n and n.get('projectValue')}")

    real = Path(r"d:\Downloads\Rajkumar Ji.xlsx")
    if real.is_file():
        live = parse_upload(real.name, real.read_bytes())
        _ok(live["advanceCount"] >= 13, f"Rajkumar advances >=13 got {live['advanceCount']}")
        _ok(abs(live["advanceTotal"] - 1695000) < 1, f"Rajkumar advance 1695000 got {live['advanceTotal']}")
        _ok(live["quoteCount"] >= 6, f"Rajkumar stages >=6 got {live['quoteCount']} {live.get('stages')}")
        _ok(live["projectValue"] > 2_500_000, f"Rajkumar value > 25L got {live['projectValue']}")
        print("Rajkumar stages:", live.get("stages"))
        print("Rajkumar value/adv/bal:", live.get("projectValue"), live.get("advanceTotal"), live.get("balance"))
    else:
        print("SKIP live Rajkumar xlsx (not on disk)")

    tmp = Path(tempfile.mkdtemp(prefix="weos_imp_"))
    os.environ["WEOS_DATA_DIR"] = str(tmp / "data")
    os.environ.pop("DATABASE_URL", None)
    os.environ.pop("WEOS_DATABASE_URL", None)
    os.environ.pop("POSTGRES_URL", None)
    from WEOS.db.engine import init_db
    from WEOS.factory.project_import import commit_imported_project
    from WEOS.factory.project_store import load_project
    from WEOS.factory.ledger_store import list_advances_for_account

    init_db()
    saved = commit_imported_project(
        pack,
        customer="Demo ji",
        customer_mobile="9999999999",
        customer_address="Hyderabad",
        project_name="Demo import job",
        import_advances=True,
    )
    _ok(saved.get("ok") and saved.get("quoteCount") == 2, f"commit quotes {saved}")
    _ok(saved.get("advanceCount") == 2, f"commit advances {saved.get('advanceCount')}")
    doc = load_project(saved["project"]["projectId"])
    _ok(len(doc.get("packageQuotes") or []) == 2, "saved package quotes")
    _ok(any(str(e.get("action")) == "import_project_files" for e in (doc.get("revisionLog") or [])), "revision log kept internally")
    _ok("revisionLog" in doc, "edit record on project")
    advs = list_advances_for_account(names=["Demo ji"])
    _ok(len(advs) >= 2, f"ledger advances {len(advs)}")
    _ok(all(a.get("id") for a in advs), "each advance has id so slip can be generated later")

    again = commit_imported_project(
        pack,
        customer="Demo ji",
        customer_mobile="9999999999",
        project_id=saved["project"]["projectId"],
        import_advances=True,
    )
    _ok(again.get("skippedCount") == 2, f"re-import skips both stages got skipped={again.get('skippedCount')} added={again.get('addedCount')} {again}")
    _ok(again.get("addedCount") == 0, f"re-import adds none got {again.get('addedCount')}")
    _ok(again.get("quoteCount") == 2, f"re-import must not double quotes got {again.get('quoteCount')}")
    _ok(again.get("advanceSkipped") == 2, f"re-import skips advances got {again.get('advanceSkipped')} added={again.get('advanceCount')}")
    doc2 = load_project(saved["project"]["projectId"])
    _ok(len(doc2.get("packageQuotes") or []) == 2, f"still 2 quotes after re-import got {len(doc2.get('packageQuotes') or [])}")

    extra = {
        "quotes": (pack.get("quotes") or [])
        + [
            {
                "note": "ACP cladding extra sheet",
                "sheetName": "ACP - extra",
                "gstMode": "exclude",
                "gstPercent": 18,
                "items": [{"category": "other", "amount": 50000, "unit": "sft", "note": "ACP new sheet"}],
            }
        ],
        "advances": [],
        "sources": ["demo-extra.xlsx"],
        "customerHint": {"name": "Demo ji"},
    }
    third = commit_imported_project(
        extra,
        customer="Demo ji",
        customer_mobile="9999999999",
        project_id=saved["project"]["projectId"],
        import_advances=False,
    )
    _ok(third.get("addedCount") == 1, f"new sheet added got added={third.get('addedCount')} skipped={third.get('skippedCount')}")
    _ok(third.get("skippedCount") == 2, f"old sheets skipped got {third.get('skippedCount')}")
    _ok(third.get("quoteCount") == 3, f"3 quotes after new sheet got {third.get('quoteCount')}")

    twins = [
        {"id": "pq_dup_a", "note": "29MM luxury", "sheetName": "29MM luxury", "items": [{"category": "window", "amount": 46000, "note": "SR sliding"}]},
        {"id": "pq_dup_b", "note": "29MM luxury", "sheetName": "29MM luxury", "items": [{"category": "window", "amount": 46000, "note": "SR sliding"}]},
        {"id": "pq_ok", "note": "Grills", "sheetName": "Grills", "items": [{"category": "grill", "amount": 12000, "note": "GF grill"}]},
    ]
    from WEOS.factory.package_quote import collapse_duplicate_package_quotes, normalize_package_quotes

    collapsed = collapse_duplicate_package_quotes(normalize_package_quotes(twins))
    _ok(len(collapsed) == 2, f"collapse duplicate stages to 2 got {len(collapsed)}")

    print("ALL OK")


if __name__ == "__main__":
    main()
